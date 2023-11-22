# Info from computer 
from os.path import basename, normpath, dirname
from os import listdir

# Packages from signal processing 
from scipy import signal
import h5py 
from scipy.stats import sem

# For gui/plots
import tkinter as tk 
import matplotlib.pyplot as plt 
import matplotlib.widgets as wdg
import matplotlib.patches as pt 
from mpl_toolkits.axes_grid1 import make_axes_locatable
from PIL import Image

# For text
import re
import string

# Math
import numpy as np 
from random import randint
from math import isnan 
import scipy.stats as stats
from math import log10


# Reading data 
import openpyxl as op 
import h5py
import pickle

from general_parameters import * 
from specific_parameters import *

# We ask the user which parameters to use 
print('enter name')
spe = get_param(input())

## Getting started 

def fill_info_file() : 
    '''
    This function reads the info file from parameters_file and let us fill the information with a GUI and put them after the last row 
    No input, no output
    '''
    paths = choose_files_from_folder() # We give a folder and there will be a list of files from which we can choose the one we want 
    dico = {}

    for n,path in enumerate(paths) :

        # Open the info file 
        wb = op.load_workbook(spe.infos) 
        sheet = wb['conditions'] # Go to the right sheet 
        row = sheet.max_row + 1 # Start writing after last row to avoid overwrite 

        date_time = date_time_from_path(path)

        # Using the GUI, the user can enter the right information for each relevant info (see parameters_file: info_keys)
        if n == 0 : 
            for key in infos_keys[1:] :
                dico[key] = {'label' : f'For file {date_time}, enter {key}', 'var': [], 'value':''}
            
        else : 
            for key in infos_keys[1:] :
                dico[key]['label'] = f'For file {date_time}, enter {key}'

        dico = make_choices(dico, text_keys = infos_keys[1:], checkbox_keys=[])

        # We fill the information in the excel file 
        column = 1 
        sheet.cell(row=row, column=column).value = path
        for key in infos_keys[1:] : 
            column += 1 
            sheet.cell(row=row, column=column).value = dico[key]['value']

        # Done, the file is saved 
        wb.save(spe.infos)
    
    expand_ID(start = row-n)


def show_image_tag_channels(date_time: string, channels: list) : 
    '''
    Input: date_time the date and the time of the recording of interest 
    Channels : list of channels to tag 
    '''

    # We start by finding the experimental ID for the slice in the info file 
    wb = op.load_workbook(spe.infos)
    sheet = wb['conditions']
    dt = date_time.replace('_','T')
    col_layout = infos_keys.index('layout') + 1
    col_ID = infos_keys.index('ID') + 1
    
    for row in range (2,sheet.max_row + 1) : 
        path = sheet.cell(row=row, column=1).value
        if path is not None and dt in path :
            row_interest = row 
            exp_ID = sheet.cell(row=row_interest, column=col_ID).value 

    # If the positions already exist for this experimental ID we stop 
    try :
        p = read_dict(f'{spe.positions}{exp_ID}.pkl')
        print('positions already found')
        return ()
    
    except :
        pass 


    def position_channels(channels: list, ref: list, distances:dict) : 
        '''
        Inputs: list of channels 
                ref : [name of ref channel, its position]
                distances {letter : the distance (x,y) between a channel and the next channel with the same number but different letter,
                        number : the distance (x,y) between a channel and the next channel with the same letter but different number}

        Outputs: x_channels with the x of all the channels in the same order as input channels (list of floats)
                y_channels with the y of all the channels in the same order as input channels (list of floats)

        '''
        # List of letters in the alphabet without the I (no I in the channels' names)
        alphabet = list(string.ascii_uppercase)
        i = alphabet.index('I')
        del alphabet[i]

        # The reference channel is the upper left of the center channels (indice 0)
        ref_letter_index = alphabet.index(ref[0][0]) # We find its position in the alphabet (B would be second etc)
        ref_number = int(ref[0][1:]) 

        # Initialize variables for output 
        x_channels = []
        y_channels = []

        # For each channel 
        for channel in channels :
            diff_letter = alphabet.index(channel[0]) - ref_letter_index # There are X letters between this channel and the ref 
            diff_number = int(channel[1:]) - ref_number # There are X numbers between this channel and the ref 
            pos = ref[1] + distances['letter']*diff_letter + distances['number']*diff_number # We can compute its position 

            x_channels.append(pos[0]) # add x to the list 
            y_channels.append(pos[1]) # add y to the list 

        return x_channels, y_channels

    
    # Get the path of the image based on the date and time of the experiment 
    image = get_image_from_info(date_time)
    print(image)

    # What we ask the user to tag 
    center_channels = 'center up left', 'center up right', 'center bottom left', 'center bottom right'

    # Convert the image into an array of RGB values 
    img = np.asarray(Image.open(image))

    # Plot the image 
    fig, ax = plt.subplots(figsize=(20,20))
    fig.set_facecolor('black')
    ax.imshow(img)
    plt.ion()

    # We add borders around the figure where we will put text later on 
    xlim = np.array(ax.get_xlim()) + [-border,+border]
    ax.set_xlim(xlim)

    ylim = np.array(ax.get_ylim()) + [-border,+border]
    ax.set_ylim(ylim)

    # The position of the text in the image 
    xtextleft = xlim[0] - text_space
    xtextright = xlim[1] + text_space
    ytextup = ylim[1] + text_space

    # In parameters file we give two types of layout and each has different center channels, the keys are the names of the layout 
    keys = list(center_channels_options.keys())


    # We ask the user to click on the first channel of interest 
    plt.text(xtextleft, ytextup, f'Click on {center_channels[0]}', size=15, 
            ha="center", va="center",
            bbox=dict(boxstyle="round",
                    ec=edgecolor,
                    fc=facecolor,
                    ))
    

    
    # Need to create a class to save the values with which we interact using a GUI 
    class define_grid :
        points = [] 
        i = 0
        x_c = {}
        y_c = {}
        tokeep = None
        pause = True 

        def mouse_event(self, event):
            # We get the x,y coordinates of the mouse when clicking 
            self.x = event.xdata
            self.y = event.ydata


            # If we have not yet clicked on all the 4 channels needed:
            if self.i < len(center_channels) : 

                
                plt.scatter(self.x, self.y, color = 'white') # We put a white dot on the last clicked channel 
                self.points.append(np.array([self.x,self.y])) # We save the coordinates of this channel 
                self.i += 1 # We count that we have saved one more channel 

                if self.i < len(center_channels) : # Do we still have more channels to click on? 
                    # If yes, we put on the screen what we should click on next 
                    plt.text(xtextleft,ytextup,f'Click on {center_channels[self.i]}', size=15, 
                    ha="center", va="center",
                    bbox=dict(boxstyle="round",
                            ec= edgecolor,
                            fc= facecolor,
                            ))

                else : 
                    # If there are no more channels to click on, we let the user decide which layout they want to try 

                    # Layout 1 click on the left side of the screen 
                    plt.text(xtextleft, ytextup,f'Click here for {keys[0]} layout', size=15, 
                        ha="center", va="center",
                        bbox=dict(boxstyle="round",
                                ec=edgecolor,
                                fc=facecolor,
                                ))
                    
                    # Layout 2 click on the right side of the screen 
                    plt.text(xtextright, ytextup, f'Click here for {keys[1]} layout', size=15, 
                        ha="center", va="center",
                        bbox=dict(boxstyle="round",
                                ec=edgecolor,
                                fc=facecolor,
                                ))

                    # We compute the distances between the channels we clicked on, whats the distance between two channels with different letters? Numbers?
                    # Since we have four points, we have two distances for each case and take the mean of the two
                    self.distances = {'letter' : np.mean([self.points[1]-self.points[0], self.points[3]-self.points[2]], axis=0), 'number' : np.mean([self.points[2]-self.points[0], self.points[3]-self.points[1]], axis=0)}

                    # For each layout, we find out where would be the channels on the image using the function 'position_channels' (go to this function for more info)
                    for c in range (2) : 
                        self.ref = [center_channels_options[keys[c]][0], self.points[0]]
                        xi,yi = position_channels(channels[keys[c]], self.ref, self.distances)
                        self.x_c[keys[c]] = xi
                        self.y_c[keys[c]] = yi 

                    self.i += 1
                
               
            else :
                if abs(self.x - xtextleft) < abs(self.x - xtextright) :
                    show = 0
                    hide = 1
                
                # If our click is closer to the second option, we will show the second layout (1) and hide the first layout (0)
                else :
                    show = 1
                    hide = 0 

                # We put first a blank n_channels on the channels of the layout to hide 
                for xi,yi in zip(self.x_c[keys[hide]], self.y_c[keys[hide]]) : 
                    plt.text(xi,yi,'     ', size=12, 
                        ha="center", va="center",
                        bbox=dict(boxstyle="round",
                                ec='black',
                                fc='black',
                                ))
                    
                # Then we can show the right layout 
                for xi,yi,chan in zip(self.x_c[keys[show]], self.y_c[keys[show]], channels[keys[show]]) : 
                    plt.text(xi,yi,chan, size=12, 
                        ha="center", va="center",
                        bbox=dict(boxstyle="round",
                                ec='white',
                                fc='white',
                                ))
                
                # We have to remember which layout was the last clicked on  (it will be the one saved for later)
                self.tokeep = keys[show]

            # We have to show what changes we did 
            fig.canvas.draw_idle()


        def on_close(self, event):
            self.pause = False 
            
            # We put the type of layout in the info file at the right row in the right column 
            sheet.cell(row=row_interest, column=col_layout).value = self.tokeep 

            # We save the file 
            wb.save(spe.infos)
            print(f'Layout of {image} saved in {spe.infos}')


            # The positions corresponding to the right layout are saved  
            # 
            positions_dict = {}

            # For each channel, we put its x,y coordinates from the right layout (tokeep)
            for ci, xi, yi in zip(channels[self.tokeep], self.x_c[self.tokeep], self.y_c[self.tokeep]) : 
                positions_dict[ci] = (xi,yi)

            # We save the file 
            save_dict(positions_dict, f'{spe.positions}{exp_ID}.pkl')
            
            print(f'Channels positions for {image} saved in {spe.positions}{date_time}.pkl')

            
    
         
    # The class needs to be called to be used 
    grid = define_grid()


    # Activate the functions depending on click 
    fig.canvas.mpl_connect('button_press_event', grid.mouse_event)
    fig.canvas.mpl_connect('close_event', grid.on_close)

    while grid.pause :
        plt.pause(0.1)
        
    plt.show()
    return
    

# Analysing the data 

def raw_signals_from_file(path:string) :
    '''
    !!!!!!! ATTENTION returns data without converting to µV !!!!!!!
    This functions is used to read a h5 file (exported with Data Manager)  

    Inputs:
    file (str): path of the h5 file 

    Output:
    raw (array): matrix with the raw data. Shape is number_channels * time_points (duration in seconds*acquisition frequency)
    channels: list of channels in this recording 
    '''

    # Get data from h5 file 
    date_time = date_time_from_path(path)

    if date_time in spe.specials.keys() :
        path = spe.specials[date_time]['path']

    try :
        data = h5py.File(path)

    except :
        print(f'{path} not found')
        return [], []

    # Find the right stream
    k = 0 
    while True :
        try :
            data = data['Data']['Recording_0']['AnalogStream'][f'Stream_{k}']
            break
        except :
            k += 1

    # Get the channels names
    all_channels = data['InfoChannel'] 
    c_index = 4 # the channel name is always at index 4 
    
    channels = [c[c_index].decode("utf-8") for c in all_channels] # convert bytes to string 
    
    # We get the data from all channels 
    data = data['ChannelData']

    raw = np.empty(data.shape, dtype = np.int32)
    data.read_direct(raw)

    return raw, channels


class preprocessing :
    def __init__(self, raw:any, peak_types:list, denoise = False, filt = None, fact_threshold = None) :
        '''
        This class creates an object that is called 'self' within the class. The object will contain the raw data, filtered signal and absolute signal
        
        Inputs :
        raw (array): raw data from ONE channel

        Output :
        self (object): object with the raw data, (denoised signal), filtered signal and absolute signal.
        '''
        self.raw = raw
        self.peak_types = peak_types
        self.duration_s = len(self.raw)/freqs

        self.signal = {'filtered' : {}, 'absolute' : {}}
        self.threshold = {}

        # Based on previous test, denoising does not give better results (need skimage if uncomment)
        # if denoise :
        #     denoised = denoise_wavelet(self.raw, wavelet = denoising_param['wavelet'], mode = denoising_param['mode'], wavelet_levels = denoising_param['wavelet_levels'])
        #     denoised = denoised * (2**31 - 1) # convert int32 to float 

        # We go through the peak types (interictal and MUA) - They need different filtering 
        for peak_type in self.peak_types : 

            if filt is not None : 
                filter_type, filter_freq = filt[peak_type]
                b,a = signal.butter(param[peak_type]['filter_order'], filter_freq, btype = filter_type, fs = freqs)
            else : 
                # Filtering with defined parameters 
                b,a = signal.butter(param[peak_type]['filter_order'], param[peak_type]['filter_freq'], btype = param[peak_type]['filter_type'], fs = freqs)
                

            # Normalization of the filtered data: abs((X - mean(X))/std(X))
            if denoise : 
                self.signal['filtered'][peak_type] = signal.filtfilt(b, a, denoised) 
            else :
                self.signal['filtered'][peak_type] = signal.filtfilt(b, a, self.raw) 

            self.signal['absolute'][peak_type] = np.abs(self.signal['filtered'][peak_type])
            
            # Apply thresholding function to find the right thresholds for this data set 
            if callable(param[peak_type]['threshold']):
                if fact_threshold is not None :
                    self.threshold[peak_type] = fact_threshold[peak_type]*param[peak_type]['threshold'](self.signal['absolute'][peak_type])
                else :
                    self.threshold[peak_type] = param[peak_type]['threshold'](self.signal['absolute'][peak_type])
             # Get threshold from parameter 
            else : 
                self.threshold[peak_type] = param[peak_type]['threshold']


class find_events :
    def __init__(self, data: object) : 
        '''
        This class creates an object that is called 'self' within the class. The object will contain the peaks' heights and time points + the mean amplitude of the peaks and their frequency

        Inputs:
        data (object): object created in preprocessing where we can find the raw data of one channel, but also the absolute data, etc
        type (str): 'interictal' or 'MUA' 
        threshold (int): a minimum value for peak detection (in mV)

        Output:
        self (object): object with the peaks' heights and time points + the mean amplitude of the peaks and their frequency
        '''

        self.total_frames = len(data.raw)
        self.thresholds = data.threshold
        self.parameters = param

        # Create empty dictionnaries to store information for each peak type 
        self.frame_index = {}
        self.amplitude = {}
        self.power = {}

        # We go through the peak types (interictal and MUA)
        for peak_type in data.peak_types : 
            self.amplitude[peak_type] = {}
            # Use the scipy function 'find_peaks' with defined parameters 
            self.frame_index[peak_type], properties = signal.find_peaks(data.signal['absolute'][peak_type], height = data.threshold[peak_type], distance = param[peak_type]['inter_event'], width = param[peak_type]['peak_duration'])

            # if at least one peak was found, find amplitude of each peak in the raw data and the absolute data 
            if len(self.frame_index[peak_type]) > 0 :
                self.amplitude[peak_type]['raw'] = [data.raw[index] for index in self.frame_index[peak_type]] # in mV 
                self.amplitude[peak_type]['absolute'] = properties['peak_heights']
                self.amplitude[peak_type]['filtered'] = [data.signal['filtered'][peak_type][index] for index in self.frame_index[peak_type]]
                self.power[peak_type] = [np.nanmean(data.raw[int(index-timepower):int(index+timepower)]**2) for index in self.frame_index[peak_type]]

            # No peak found, all variables = 0 
            else : 
                self.amplitude[peak_type]['raw'] = 0 # in mV 
                self.amplitude[peak_type]['filtered'] = 0 # in mV 
                self.amplitude[peak_type]['absolute'] = 0 # in mV 
                self.power[peak_type] = 0 # in mV
        
        


def raw_data_all_channels(path, frames, to_save = True, to_return = False) :
    '''
    This function is used to return (or save in a picke file) the raw data from all the channels between frames[0] and frames[1]
    Input
    path: path of the h5 file where the data is 
    frames: list with the [frame start, frame stop]
    to save: True/False; if True will save to a pickle file in the figure folder named with the date_time of the file and 'raw_channels' 
    to return: True/False; if True will return a dictionary with raw[number][letter] will contain the raw data between the frames of interest 
    '''

    # Need all the letters in order without the letter i (bc no 'i' in the nomenclature of MEA channels)
    alphabet = list(string.ascii_uppercase)
    i = alphabet.index('I')
    del alphabet[i]

    # Get the raw data and the list of channels 
    raw, channels_list = raw_signals_from_file(path)
    channels = ', '.join(channels_list)

    # Finds all the letters and numbers in the channels names 
    numbers = re.findall(r'\d+',channels)
    numbers = [int(number) for number in numbers]
    letters = re.findall(r'[a-zA-Z]',channels)
    min_letter = alphabet.index(min(letters))
    max_letter = alphabet.index(max(letters)) 

    # Puts the data in a dictionnary 
    raw_channels = {}
    for n in range (min(numbers), max(numbers) + 1) : 
        raw_channels[n] = {}
        for letter in range (min_letter, max_letter + 1) :
            raw_channels[n][alphabet[letter]] = 0
            channel = f'{alphabet[letter]}{n}'
            if channel in channels_list :
                channel_index = channels_list.index(channel)
                raw_channels[n][alphabet[letter]] = factor_amp*raw[channel_index][frames[0]:frames[1]]
            print(f'Done for {channel}', end = '\r')
    
    # If decided to save will save to a pickle file 
    if to_save : 
        date_time = date_time_from_path(path)
        save_dict(raw_channels, f'{spe.figures}{date_time}_raw_channels.pkl')
    
    # If decided to return will return the dict object 
    if to_return : 
        return raw_channels
    
def extract_data(path, condition, peak_types = peak_types, rewrite = False) : 
    '''
    We use this function to get the data from the h5 file, process it, extract and save the peaks for peak_types of interest based on the parameters defined in the parameters_file 
    input: path of the h5 file of interest (str) and list of peak_types (['interictal', 'MUA'] by default)
    '''
    # Call function to extract signal and channel names from the h5 file 
    date_time = date_time_from_path(path)

    if date_time in spe.specials.keys() : 
        path = spe.specials[date_time]['path']
    
    if not rewrite :
        try :
            peaks = read_dict(f'{spe.peaks}{date_time}_F6.pkl')
            return()
        except :
            pass
    
    raw, channels = raw_signals_from_file(path)

    if raw == [] :
        return()
    
    start, stop = determine_timing(date_time, condition, end = len(raw[0]))
    bad = bad_channels(date_time)

    for num_channel, channel in enumerate(channels) :
        
        if channel not in bad : 
            channel_index = channels.index(channel)
            # Take only the signal from the channel chosen
            raw_channel = raw[channel_index]
            raw_channel = raw_channel[start:stop]
            raw_channel = factor_amp*raw_channel

            # Call the class preprocessing to filter the signal 
            data = preprocessing(raw_channel, peak_types)

            # Call the class find_events to find the peaks for each defined peak type 
            peaks = find_events(data)
            peaks.date_time = date_time
            peaks.channel = channel 
            peaks.frames_used = start, stop
        

            # To save to a pickle object 
            save_dict(peaks, f'{spe.peaks}{date_time}_{channel}.pkl')

            print(f'{date_time}, {channel} ({[[len(peaks.frame_index[peak_type]),peak_type] for peak_type in peak_types]} .. {int(100*(num_channel+1)/len(channels))}% ', end = '\r')
    
    
            


# Getting / manipulating peaks 

def get_peaks_from_pkl(channel, date_time, selected_peaks = False) : 
    '''
    Pickle files contains the 'peaks' object with all the information about the peaks found for a specific recording 
    I decided to name them as '_date_time_channel.pkl' and put them in the peaks folder so they can be found easily
    Inputs :
    channel : the channel of interest
    date_time : the date and time of the recording 
    selected_peaks : if we need to take the pickle object with only a sample of the peaks 
    Outputs : 
    peaks object with self.total_frames (int), self.thresholds (dict)), self.paramameters (dict), self.frame_index (dict), self.amplitude (dict), self.power (dict)
    '''

    peaks = []

    if selected_peaks :
        name_file = f'{spe.selected_peaks}{date_time}_{channel}.pkl'
    else :
        name_file = f'{spe.peaks}{date_time}_{channel}.pkl'
    
    try : 
        peaks = read_dict(name_file) # open the pickle file 
        
    except :
        # If the pickle file is not found we create empty variables 
        print(f'no peaks found for {date_time}, {channel}')
        peaks = None 

    return peaks


def save_selected_peaks_pkl(date_time, channel, peaks, deleted) : 
    '''
    CAREFUL this deletes the old version of the peaks file 
    This functions deletes some peaks from the peaks object and save this new version 
    Inputs : 
    date_time : date time of the recording
    channel: the channel of interest 
    peaks: the peaks object 
    deleted: list of frames to delete 

    No return, saves the new pickle 
    '''
    for peak_type in peak_types : 
        if len(peaks.frame_index[peak_type]) > 0 : 
            peaks.frame_index[peak_type] = [peaks.frame_index[peak_type][i] for i in range (len(peaks.frame_index[peak_type])) if i not in deleted[peak_type]] # copy the peaks except the ones to delete 
            peaks.power[peak_type] = [peaks.power[peak_type][i] for i in range (len(peaks.power[peak_type])) if i not in deleted[peak_type]] # same for power 
            for data_type in data_types : 
                peaks.amplitude[peak_type][data_type] = [peaks.amplitude[peak_type][data_type][i] for i in range (len(peaks.amplitude[peak_type][data_type])) if i not in deleted[peak_type]] # same for amplitude

    # saves the pickle 
    save_dict(peaks, f'{spe.selected_peaks}{date_time}_{channel}.pkl')


## GUI FUNCTIONS

def make_choices(dict_object: dict, text_keys: list, checkbox_keys: list) : 
    '''
    This functions allows to show a window where the user can put information either in text or in check boxes 

    Input 
    dict_object (dict): dictionnary where each key is a variable of interest with the different subkeys: 'label', 'var', 'value' 
            'label': information to show to the user
            'var': where we store temporary variables 
            'value': the input of the user 

    text_keys (list of str): list of the keys we use when the function is called (for text entries)

    checkbox_keys (list of str): list of the keys we use when the function is called (for checkboxes)

    Output
    dict_object (dict): the dictionnary after modifications 
    '''

    def save_entry_fields():
        '''
        Get the user input (string) and store it in the dictionnary 
        '''
        # Go through all the defined keys for text entries 
        for key in text_keys :
            # Save the user input in a dictionnary 
            dict_object[key]['value'] = dict_object[key]['var'].get()

    def save_entry_fields_box():
        '''
        Get the user input (box checked or not) and store it in the dictionnary 
        '''
        # Go through all the defined keys for check_box entries 
        for key in checkbox_keys :
            # Save the user input in a dictionnary 
            dict_object[key]['value'] = dict_object[key]['var'].get()


    # Open a new window 
    master = tk.Tk()
    
    # Place the window in the center of the screen 
    master.eval('tk::PlaceWindow . center')

    # For each key related to text entries
    for i,key in enumerate(text_keys) :

        # Show text with the label defined in the dictionnary for this key 
        tk.Label(master, text=dict_object[key]['label']).grid(row=i)

        # Save the tk enty object in the dictionnary 
        dict_object[key]['var'] = tk.Entry(master)
        dict_object[key]['var'].insert(0, dict_object[key]['value'])

        # Place this entry object at the right place in the window (each entry is a row and they are all in the column 1)
        dict_object[key]['var'].grid(row=i, column=1)
    
    col = 1 
    row = len(text_keys) + 1  
    
    
    # For each key related to check boxes entries
    for i,key in enumerate(checkbox_keys) :
        
        col, row = div(i)
        col += 2
        row += 1

        # Create a tk variable for integers 
        dict_object[key]['var']= tk.IntVar()

        # Create a check box with the label defined in the dictionnary for this key 
        c = tk.Checkbutton(master, text=dict_object[key]['label'],variable=dict_object[key]['var'], onvalue=1, offvalue=0, command=save_entry_fields_box)

        # Place this checkbox at the right place in the window (each one is a row and they are all in the column 1)
        c.grid(row=row, column = col)
                     

    # Create a button for saving choices (text entries)    
    tk.Button(master, text='Save your choices', command=save_entry_fields).grid(row=row+1, column=col, sticky=tk.W, pady=4)

    tk.mainloop()
    return dict_object 


def make_it_interactive(fig: object, axes: object, param: dict) :
    '''
    Function to add sliding bars to zoom on y axis and scroll through x axis; 
    also buttons to move from one peak to another and delete it if needed

    Inputs 
    fig: matplotlib figure to modify
    axes: matplotlib axes to modify 
    param (dict): dictionnaries with the keys:
            'xlim' (list of 2 int): [min_x, max_x]
            'xlabel' (str): label for x axis 
            'ylim' (list of 2 int): [min_y, max_y]
            'ylabel' (str): label for y axis 
            'x_window'(int) : time scale for plot
            'ticks' (list of int): list of points that need to be highlighted on x axis
            'y_zoom' (list of 2 int): min and max zoom on y axis 

    Output
    peaks_to_del (list of int): indexes of peaks to delete 
    '''
    # Adjust figsize to make room for sliders
    fig.subplots_adjust(left=0.25, bottom=0.25)
    plt.ion()

    # Position each slider ([x,y,w,h] x = 0 is left and y = 0 is bottom, width goes right and height goes up)
    ax_slide_x = fig.add_axes([0.25, 0.15, 0.65, 0.03])
    ax_slide_y = fig.add_axes([0.15, 0.25, 0.03, 0.65])

    # Adding ticks if they are defined 
    ax_slide_x.add_artist(ax_slide_x.xaxis)
    ax_slide_x.set_xticks(param['ticks'])
    ax_slide_x.set_xticklabels(["*"]*len(param['ticks'])) # Shown as stars on the plot 
    
    # Create the slider objects based on the defined parameters for label, min and max values. 
    slider_x = wdg.Slider(ax=ax_slide_x, label=param['xlabel'], 
                                            valmin=param['xlim'][0], valmax=param['xlim'][1]- param['x_window'], 
                                            valinit=param['xlim'][0], orientation='horizontal')

    slider_y = wdg.Slider(ax=ax_slide_y, label=param['ylabel'], 
                                            valmin=param['y_zoom'][0], valmax=param['y_zoom'][1], 
                                            valinit=1, orientation='vertical')

    # Create buttons at the right position with labels
    # Previous button on the bottom left 
    ax_button_prev = fig.add_axes([0.25, 0.000001, 0.1, 0.08])
    prev = wdg.Button(ax_button_prev, 'Previous', color= 'grey', hovercolor='blue')

    # Delete button in the middle 
    ax_button_del = fig.add_axes([0.55, 0.000001, 0.1, 0.08])
    delete = wdg.Button(ax_button_del, 'Delete this peak', color= 'grey', hovercolor='red')

    # Next button on the bottom right 
    ax_button_next = fig.add_axes([0.85, 0.000001, 0.1, 0.08])
    next_ev = wdg.Button(ax_button_next, 'Next', color= 'grey', hovercolor='blue')
    

    # The function to be called anytime a slider's value changes
    def update_x(val):
        for ax in axes:
            ax.set_xlim([int(slider_x.val), int(slider_x.val + param['x_window'])]) # scroll through the horizontal axis based on x_window step 
        fig.canvas.draw_idle()

    def update_y(val):
        for s,ax in enumerate(axes) :
            ax.set_ylim([int(param['ylim'][s][0]*(slider_y.val)), int(param['ylim'][s][1]*slider_y.val)]) # Zoom on y axis 
        fig.canvas.draw_idle()

    # Need to create a class to update variables with button presses 
    class IndexPeak:
        # Start at i = -1 so when we click 'next' we go to the next peak (i=0)
        i_peak = -1
        peaks_to_del = []
        colors_ticks = ['black']*len(param['ticks'])
        pause = True

        # Function to go to the next peak 
        def nextone(self, event):
            if self.i_peak < len(param['ticks']) - 1 : # Move to the next peak if we are not at the last one 
                self.i_peak += 1 
            else : # Go back to the first one if we reached the end 
                self.i_peak = 0 
            slider_x.set_val(param['ticks'][self.i_peak] - int((param['x_window'])/2))
            fig.canvas.draw_idle()
            
        # Function to go to the previous peak 
        def previousone(self, event):
            if self.i_peak > 0: # Move to the peak before if we are not at the first one
                self.i_peak -= 1 
            else : # Go to the last one if we reached the beginning
                self.i_peak = len(param['ticks']) - 1 
            slider_x.set_val(param['ticks'][self.i_peak] - int((param['x_window'])/2))
            fig.canvas.draw_idle()

        # Function to delete the peak showed on the screen
        def deletethisone(self, event) : 
            self.peaks_to_del.append(self.i_peak) # We add the index of this peak to a list of peaks to delete 
            self.colors_ticks[self.i_peak] = 'red' # We put a red tick for this one to show that it has been deleted 
            for ticklabel, tickcolor in zip(ax_slide_x.get_xticklabels(), self.colors_ticks):
                ticklabel.set_color(tickcolor)
            if self.i_peak < len(param['ticks']) - 1 : # Move to the next peak if we are not at the last one 
                self.i_peak += 1 
            else : # Go back to the first one if we reached the end 
                self.i_peak = 0 
            slider_x.set_val(param['ticks'][self.i_peak] - int((param['x_window'])/2))
            fig.canvas.draw_idle()

        # Function to return the list of peaks to delete 
        def on_close(self, event) : 
            self.pause = False 


    # Call the function for each update 
    slider_x.on_changed(update_x)
    slider_y.on_changed(update_y)

    callback = IndexPeak()
    
    prev.on_clicked(callback.previousone)
    next_ev.on_clicked(callback.nextone)
    delete.on_clicked(callback.deletethisone)
    fig.canvas.mpl_connect('close_event', callback.on_close)

    while callback.pause :
        plt.pause(2)


    # Show the plot 
    plt.show()
    
    # return the updated figure with sliding bars 
    return callback.peaks_to_del


def choose_files_from_folder() :
    '''
    Ask the user to fill in the path of a folder with h5 files to look at 
    And then shows the list of h5 files in this folder so we can select which we want 

    No input

    Ouputs the list of paths chosen by the user 
    '''
     # Ask the user what is the folder with the recordings 
    gui_dict = {'Path' : {'label' : 'Path of the H5 folder', 'var' : [],'value' : ''}}
    gui_dict = make_choices(gui_dict, text_keys = ['Path'], checkbox_keys = [])
    
    # Save the file path in r format to avoid problems with '\'
    folder = r'{}'.format(gui_dict['Path']['value'])
    files = listdir(folder)
    files = [r'{}'.format(file) for file in files if '.h5' in file]

    # No h5 files in the list 
    if len(files) == 0 : 
        print('Wrong folder? No h5 files found here')

    # For readibility, we show only the date_time of the recording to the user so he/she can chooses 
    for file in files :
        date_time = file[:file.index('Mcs')].replace('T', '_')
        gui_dict[file] = {'label' : date_time, 'var' : [],'value' : 0}
    gui_dict = make_choices(gui_dict, text_keys = [], checkbox_keys = files) # GUI

    # All the checkbox clicked on will have a '1' value so we can save the path selected by the user 
    paths = []
    for file in files :
        if gui_dict[file]['value'] == 1 :
            paths.append(folder + '\\' + file )

    return paths

def choose_plot_from_folder() :
    '''
    Looks at the files in the param folder 
    And then shows the list of plots parameters so we can select which we want 

    No input

    Ouputs the list of plots chosen
    '''

    # Lists the files in the param folder
    files = listdir(spe.param)
    files = [r'{}'.format(file) for file in files if '.pkl' in file]

    # No pkl files in the list 
    if len(files) == 0 : 
        print('Wrong folder? No pkl files found here')

    gui_dict = {}
    # For readibility, we show only the date_time of the recording to the user so he/she can chooses 
    for file in files :
        gui_dict[file] = {'label' : file, 'var' : [],'value' : 0}
    gui_dict = make_choices(gui_dict, text_keys = [], checkbox_keys = files) # GUI

    # All the checkbox clicked on will have a '1' value so we can save the path selected by the user 
    plots = []
    for file in files :
        if gui_dict[file]['value'] == 1 :
            plots.append(file.replace('.pkl', ''))

    return plots

def choose_files_from_info(condition = False) :
    '''
    Get the list from file in the info file and the user can click on the ones of interest
    No input
    Ouput the list of paths selected by the user 
    '''

    # Find files from info file 
    if condition :
        paths_info, _, conditions = paths_channels_from_info(condition=True)
    else : 
        paths_info, _ = paths_channels_from_info()

    # Save the file path in r format to avoid problems with '\'
    gui_dict = {}

    # For readibility, we show only the date_time of the recording to the user so they can choose
    for path in paths_info :
        date_time = date_time_from_path(path)
        gui_dict[path] = {'label' : date_time, 'var' : [],'value' : 0}

    gui_dict['all'] = {'label' : "select all", 'var' : [],'value' : 0}

    box_keys = paths_info + ['all']
    gui_dict = make_choices(gui_dict, text_keys = [], checkbox_keys = box_keys) # GUI 

    # All the checkbox clicked on will have a '1' value so we can save the path selected by the user 
    
    if gui_dict['all']['value'] == 0 :
        paths = []
        for path in paths_info :
            if gui_dict[path]['value'] == 1 :
                paths.append(path)
    else : 
        paths = paths_info

    if condition : 
        return paths, conditions
    else : 
        return paths 


    

# Manipulating channels and their positions 

def listing_channels(to_print = True, to_return = False, specific_layout = None) :
    '''
    This function prints the name of channels in the 'sparse' and/or the 'dense' layout (both have 120 channels)
    The sparse layout is a square, columns from A to M and rows from 1 to 10 
    The dense layout is more circular, columns from A to M and rows from 1 to 12 but not all combinations are represented 
    '''
    list_channels = {}
    # List of letters in the alphabet without the I (no I in the channels' names)
    alphabet = list(string.ascii_uppercase)
    i = alphabet.index('I')
    del alphabet[i]
    alphabet = alphabet[:alphabet.index('N')] # We go only to M (last column in both layouts)
    list_channels['sparse'] = [f'{letter}{number}' for letter in alphabet for number in range (1,11)] # sparse is a square
    todel = [f'{l}{i}' for i in range (1,4) for l in ['A', 'M']]
    todel += [f'{l}{i}' for i in range (1,3) for l in ['B', 'L']]
    todel += [f'{l}1' for l in ['C', 'K']]
    todel += [f'{l}{i}' for i in range (10,13) for l in ['A', 'M']]
    todel += [f'{l}{i}' for i in range (11,13) for l in ['B', 'L']]
    todel += [f'{l}12' for l in ['C', 'K']]
    list_channels['dense'] = [f'{letter}{number}' for letter in alphabet for number in range (1,13) if f'{letter}{number}' not in todel] # dense layout goes up to 12 but we delete some combinations 
    
    if specific_layout is not None : 
        list_channels = list_channels[specific_layout] 
    if to_print : 
        print(list_channels)
    if to_return : 
        return (list_channels)


def get_channels_positions(date_time: string) : 
    '''
    Inputs: date_time date and time of the file of interest 
        
    Outputs channels list complete
            x coordinates of the channels
            y coordinates of the channels 

    '''

    experiments = experiments_dict(return_exp_only=True)
    exp_ID = experiments[date_time]['ID']

    # open the file
    positions = read_dict(f'{spe.positions}{exp_ID}.pkl')
    
    channels =[]
    x = []
    y =[]
    # Get the x,y values from the file 
    for channel in positions.keys() : 
        channels.append(channel)
        x.append(positions[channel][0])
        y.append(positions[channel][1])

    return channels, x,y


# Getting info from files 

def paths_channels_from_info(condition = False, start_row = start_row) : 
    '''
    Starting at the start_row defined in the parameter file we get the paths and the list of channels that need to be analysed 
    No inputs
    Output : list of paths of interest and dictionnary with paths as keys and channels of interest for each 
    '''

    # Read the info file at the right sheet 
    wb = op.load_workbook(spe.infos)
    sheet = wb['conditions']

    # Initialize variables 
    paths = []
    channels_to_keep_by_path = {}
    conditions_by_path = {}

    path_column = infos_keys.index('path') + 1
    layout_column = infos_keys.index('layout') + 1
    condition_column = infos_keys.index('condition') + 1

    # Get the paths in the first column and the channels in the second column starting from start_row 
    for row in range (start_row, sheet.max_row + 1) : 
        path = sheet.cell(row=row, column=path_column).value
        
        if path is not None:
            paths.append(r'{}'.format(path))
            layout = sheet.cell(row=row, column=layout_column).value
            channels_to_keep_by_path[path] = listing_channels(to_print=False, to_return=True, specific_layout=layout)
            if condition :
                conditions_by_path[path] = sheet.cell(row=row, column=condition_column).value

    # Save the file 
    wb.save(spe.infos)
    
    if condition : 
        return paths, channels_to_keep_by_path, conditions_by_path
    else :
        return paths, channels_to_keep_by_path




def get_image_from_info(date_time: string) : 
    '''
    Input: date_time the date_time of the file of interest 
    Output: the path where the image of the slice recorded at this date_time 
    '''

    # Reads the info file at the right sheet 
    wb = op.load_workbook(spe.infos)
    sheet = wb['conditions']

    date_time = date_time.replace('_','T')

    # Look for the right date and time and get the number of the slice in the right column 
    path_column = infos_keys.index('path') + 1
    slice_column = infos_keys.index('slice') + 1

    for row in range (start_row, sheet.max_row + 1) : 
        path_found = sheet.cell(row=row, column=path_column).value
        if type(path_found) == str and date_time in path_found:
            n = sheet.cell(row=row, column=slice_column).value
            path = path_found

    # Save the file 
    wb.save(spe.infos)

    folder_img = dirname(path)
    
    for f in format_slices :
        image = r'{}'.format(f'{folder_img}\slice{n}.{f}')
        try :
            _ = Image.open(image)
            return image
        except : 
            pass

        
    print(f'{folder_img}\\slice{n} not found')

def date_time_from_path(path:string) -> string: 
    '''
    Input: path the path to a h5 file
    Output: date_time as YYYY-MM-DD_HH-MM-SS
    '''
    namefile = basename(normpath(path)) # Get the name of the file from the path 
    date_time = namefile[:namefile.index('Mcs')].replace('T', '_') # Only keep the characters before 'Mcs' and replace the 'T' with '_'

    return date_time

def path_from_datetime(date_time:string) -> string:
    '''
    Inputs: date_time as YYYY-MM-DD_HH-MM-SS 
    Output: path of the h5 file corresponding to date_time given 
    '''
    if date_time in spe.specials.keys() :
        return spe.specials[date_time]['path']
    paths, _ = paths_channels_from_info() # Get a list of all the paths 
    for path in paths :
        date_time = date_time.replace('_','T') # We save the date_time with an underscore but in the paths there is a 'T' instead 
        if date_time in path : # Return the path if it corresponds to the date_time given
            return path


def pickle_durations() :
    '''
    TO DO 
    '''
    try : 
        duration_in_frames = read_dict( f'{spe.results}durations.pkl')
    except : 
        duration_in_frames = {}
    paths, _ = paths_channels_from_info()
    for path in paths : 
        date_time = date_time_from_path(path)
        print(f'{date_time} processing', end = '\r')
        if date_time not in duration_in_frames.keys() : 
            duration_in_frames[date_time] = get_number_frames(path)
    save_dict(duration_in_frames, f'{spe.results}durations.pkl')

def get_number_frames(path) : 
    '''
    From a h5 file, finds the number of frames 
    LONG
    Input: path (str) 
    Output: Number of minutes of this recording
    '''
    
    raw, _ = raw_signals_from_file(path)

    # We find the length of data stored for the first channel 
    return len(raw[0])

def add_duration_to_infos() :
    '''
    Modifies the info file to add the duration in minutes in the last column based on the number of frames 
    a bit long because need to look at the h5 file (20s per file)
    '''

    # open the info file 
    wb = op.load_workbook(spe.infos)
    sheet = wb['conditions']
    maxcol = sheet.max_column + 1

    c = 1
    minutes_col = None
    while c <= maxcol : 
        if sheet.cell(row=1, column=c).value == 'minutes' :
            minutes_col = c 
        c += 1

    
    if minutes_col is None :
        print('column for minutes not found')
        minutes_col = maxcol 

    path_column = infos_keys.index('path') + 1
    # Find the path in the first column of each row 
    for row in range (start_row, sheet.max_row + 1) : 
        path = sheet.cell(row=row, column=path_column).value
        if path is not None:
            # Get the duration and write it 
            minutes = get_number_frames(path)/(60*freqs)
            sheet.cell(row=row, column=maxcol).value = minutes 
        print(f'{int(100*row/sheet.max_row)}% done', end = '\r')

    # Save the file 
    wb.save(spe.infos)
    

def bad_channels(date_time) : 
    '''
    Reads the info file to find the channels with noise for this file 

    Input : date_time = date time of the file of interest
    Output : list of bad channels 
    
    '''
    wb = op.load_workbook(spe.infos)
    sheet = wb['conditions']

    date_time = date_time.replace('_','T')

    # Finds the column with noisy channels 
    noise_col = None
    for col in range (1,sheet.max_column+1) :
        if sheet.cell(row=1, column=col).value in ['bruit','Bruit'] :
            noise_col = col
    
    path_column = infos_keys.index('path') + 1
    # If the right column was found, find the row for the file of interest and returns a list of the bad channels 
    if noise_col is not None : 
        for row in range (start_row, sheet.max_row + 1) : 
            path_found = sheet.cell(row=row, column=path_column).value
            if type(path_found) == str and date_time in path_found:
                bad = sheet.cell(row=row, column=noise_col).value
                bad = bad.replace(' ','').split(',')   # this turns a string list to a list of strings   
    else :
        print('Noise column not found')
        bad = []

    return bad 


def get_channels_and_z(date_time, condition, peak_criteria, peak_type, echo = False) : 
    '''
    This functions finds all the pkl file corresponding to the date_time of interest (meaning for all channels) and get the channel name and peak info from this 
    Input: date_time 
    peak_criteria: one of the following ['frequency', 'amplitude', 'power']
    peak_type: interictal/MUA

    Output 
    channels: list of channels analysed 
    z: the value of interest for each channel 
    channels and z are in the same order so easy to know which value corresponds to which channel and vice versa
    '''

    # List of all the files with the right date and time (different channels)
    list_peaks = list_date_time_pkl(date_time)

    channels = []
    z = []

    # For each finds the name of the channel and the information depending on the peak criteria (frequency, median amplitude or median power of the peaks for the channel)
    for num_file, file in enumerate(list_peaks) :

        channel = channel_from_pkl(date_time, file)
        channels.append(channel)
        z.append(peak_info(date_time, condition, channel, peak_type, peak_criteria))

        print(f'{int(100*num_file/len(list_peaks))}% done for {date_time}', end = '\r')
    
    return channels, z 


def determine_timing(date_time, condition, end = None) : 
    '''
    This function is used to get the first and last frame to take into account in the recording 
    --> see 'timing' in the specific parameters 

    '''

    # For some conditions 
    if date_time in spe.specials.keys() :
        begin, end = spe.specials[date_time]['start_stop']
        begin = begin*freqs
        if end is not None:
            end = end*freqs
    
    else : 
        begin = 0

    if end is None :
        try : 
            duration_in_frames = read_dict( f'{spe.results}durations.pkl')
            end = duration_in_frames[date_time]
        except :
            path = path_from_datetime(date_time) 
            end = get_number_frames(path)

    # For the 'end' type: we take n minutes before the end of the recording (before we add the drug)
    if spe.timing[condition]['type'] == 'end' : 
        stop = end - spe.timing[condition]['time']
        start = stop - spe.timing[condition]['duration'] 
        if start < 0 : 
            print(f'error duration for {date_time}, start is {begin/(freqs*60)}min, duration is {(end-begin)/(freqs*60)}min instead of {spe.timing[condition]["duration"]/(freqs*60)}')
            start = 0

    # For the 'after' type, we wait n minutes and then take the duration 
    if spe.timing[condition]['type'] == 'after' : 
        start = begin + spe.timing[condition]['time']
        stop = start + spe.timing[condition]['duration']
        if stop > end : 
            print(f'error duration for {date_time}, stop is {end/(freqs*60)}min, duration is {(end-start)/(freqs*60)}min instead of {spe.timing[condition]["duration"]/(freqs*60)}, start is {start/(freqs*60)}min')
            stop = end
    
    return int(start), int(stop )
        
def condition_to_num(condition) :
    '''
    baseline is 0, drug is 1, washout is 2 
    '''
    if condition == 'baseline_drug_washout' :
        num = 0
    elif condition == 'baseline' :
        num = 0 
    elif condition == 'washout' :
        num = 2
    else :
        num = 1
    return num 


def peak_info(date_time, condition, channel, peak_type, peak_criteria = 'amplitude', peaks = None) : 
    '''
    Input: date_time (str)
    condition (str in baseline, drug, washout)
    channel: name of the channel
    peak_criteria: one of the following ['frequency', 'amplitude', 'power']
    peak_type: interictal/MUA

    '''

    # opens the file with the peak object 
    if peaks is None:
        peaks = read_dict(f'{spe.peaks}{date_time}_{channel}.pkl')

    # look at the time points with peaks 
    frames = peaks.frame_index[peak_type]
    
    # If None then nothing
    if np.isnan(frames).all() or frames == [] : 
        indices = []

    # If the user defined a peak criteria that we do not offer, print error 
    if peak_criteria not in ['frequency', 'amplitude', 'power'] :
        print(f"ERROR criteria not found, use one of {['frequency', 'amplitude', 'power']} ")
        return 0
    
    # No points found 
    if len(frames) <= 0 : 
        return 0

    # If user wants frequency, returns it (number of peaks * acquisition freq / time_window) Hz 
    if peak_criteria == 'frequency' : 
        return len(frames)*freqs/spe.timing[condition]['duration']
    
    # For the amplitude, we take the median amplitude (we use absolute value to take into account negative peaks)
    if peak_criteria == 'amplitude' : 
        amplitudes_raw = peaks.amplitude[peak_type]['raw']
        return np.nanmedian(np.abs(amplitudes_raw))
    
    # Finally we can also return the median power 
    if peak_criteria == 'power' : 
        power = peaks.power[peak_type]
        list_powers = [10*log10(power_i*1e-12) for power_i in power] # Converts average power in µV^2 to dbW
        return np.nanmedian(list_powers)

def save_raw_data(path, channel, frames = None) : 
    '''
    This function saves the raw data of a channel into a pickle file 
    '''
    date_time = date_time_from_path(path) # finds the date time of the recording 
    print(f'processing {date_time}')
    raw, channels = raw_signals_from_file(path) # get the data for all channels 
    channel_index = channels.index(channel) # find the position of the right channel
    raw_channel = factor_amp*raw[channel_index] # converts to µV

    if frames is not None : 
        raw_channel = raw_channel[frames[0]:frames[1]] # get only the data at the frames of interest 
    
    

    save_dict(raw_channel, f'{spe.results}raw_{date_time}_{channel}.pkl') # saves to a picke file in the results folder 


### Organize our results 


def order_channels(ID_dict, peak_type, condition_ref = 'baseline', peak_criteria = {'interictal' : ['amplitude', 'frequency'], 'MUA' : 'frequency'}, save_raw = False, spe_IDs = None) : 
    '''
    Takes less than a minute for everything
    Creates a pickle object with the channels ordered from best to worse based on the criteria defined 
    '''

    def process_ID(channels_ordered, exp_ID) : 
        print(f'Processing {exp_ID}')
        if type(peak_criteria[peak_type]) == str : 
            channels, z = get_channels_and_z(ID_dict[exp_ID]['date_times'][condition_to_num(condition_ref)], condition_ref, peak_criteria[peak_type], peak_type)
        elif type(peak_criteria[peak_type]) == list : 
            channels = [[] for _ in range (len(peak_criteria[peak_type]))]
            z = [[] for _ in range (len(peak_criteria[peak_type]))]
            for c,crit in enumerate(peak_criteria[peak_type]) : 
                channels[c], z[c] = get_channels_and_z(ID_dict[exp_ID]['date_times'][condition_to_num(condition_ref)], condition_ref, crit, peak_type)
                z_mean = np.nanmean([zi for zi in z[c] if zi > 0])
                z[c] = [zi/z_mean for zi in z[c]] 
            for chan1 in channels : 
                for chan2 in channels  :
                    if chan1 != chan2 :     
                        print('ERROR')
            channels = channels[0]
            z = [np.nanmean([z[i][c] for i in range(len(z))]) for c in range (len(channels))]
        # we are only interested in the N_best (see parameters_file) channels for comparing between conditions
        index_list = list(np.argsort(z))
        index_list.reverse()
        channels_ordered[exp_ID] = [channels[i] for i in index_list]
        if save_raw : 
            for d,date_time in enumerate(ID_dict[exp_ID]['date_times']) :
                path = path_from_datetime(date_time)
                frames = determine_timing(date_time, condition =ID_dict[exp_ID]['conditions'][d])
                try : 
                    read_dict(f'{spe.results}raw_{date_time}_{channels_ordered[exp_ID][0]}.pkl')
                except : 
                    save_raw_data(path, channels_ordered[exp_ID][0], frames)
        return channels_ordered


    if spe_IDs is not None :
        channels_ordered = read_dict(f'{spe.results}{peak_type}_chan_ord.pkl')
        for exp_ID in spe_IDs : 
            channels_ordered = process_ID(channels_ordered, exp_ID)
    
    else : 
        channels_ordered = {}
        for exp_ID in ID_dict.keys() :
            print(f'Processing {exp_ID}')
            channels_ordered = process_ID(channels_ordered, exp_ID)

    save_dict(channels_ordered, f'{spe.results}{peak_type}_chan_ord.pkl')


def experiments_dict(return_exp_only = False) : 
    '''
    This function allows to get either a dictionnary with relevant info for each date_time (return_exp_only True) or for each exp_ID (return_exp_only False)
    The information is found in the spe.infos
    '''

    # Open the file with infos 
    wb = op.load_workbook(spe.infos)
    sheet = wb['conditions']
    row = 2
    experiments = {}
    path_column = infos_keys.index('path') + 1
    max_column = sheet.max_column + 1

    for row in range (start_row, sheet.max_row+1) :
        # Each row corresponds to a file with all its information (the path of the file, the ID of the experiment, the condition (baseline, etc)...)
        path = sheet.cell(row=row, column=path_column).value
        if path is not None : 
            namefile = basename(normpath(path)) # get the file name from the path 
            date_time = namefile[:namefile.index('Mcs')].replace('T', '_') # Date time is stored as YYYY-MM-DDTHH-MM-SS but we change the 'T' to '_' 
            experiments[date_time] = {}
            for column in range (path_column +1, max_column) : 
                experiments[date_time][sheet.cell(row=1, column=column).value] = sheet.cell(row=row, column= column).value # Finally we put the all the information found in the dict 

    wb.save(spe.infos) # save the file and returns the dict 

    if return_exp_only :
        return experiments # For each date_time, we can know the information of interest 
    
    # Create a dict object ordered by exp_ID 
    ID_dict = {}
    drugs_list = {}

    # Scroll through the experiments 
    for date_time in experiments.keys() :
        # Find the exp ID
        exp_ID = experiments[date_time]['ID']
        
        # Create a field for this exp ID if new and add the informations 
        if exp_ID not in ID_dict.keys() :
            ID_dict[exp_ID] = {'date_times' : [date_time], 'conditions' : [experiments[date_time]['condition']]}
            ID_dict[exp_ID]['slice'] = experiments[date_time]['slice']
            ID_dict[exp_ID]['layout'] = experiments[date_time]['layout']
        else : 
            # If the field already exists just add to this field the date time of this recording and its condition 
            ID_dict[exp_ID]['date_times'].append(date_time)
            ID_dict[exp_ID]['conditions'].append(experiments[date_time]['condition'])

        # In baseline drug washout experiments, we use different drugs so we need to know for each exp ID which drug was used
        if experiments[date_time]['condition'] not in ['baseline', 'washout'] :
            ID_dict[exp_ID]['drug'] = experiments[date_time]['condition']
            
            # We will make our analysis separately for each drug so we need to have a list of experiment IDs for each drug 
            if experiments[date_time]['condition'] in drugs_list.keys() :
                drugs_list[experiments[date_time]['condition']].append(exp_ID)
            else : 
                drugs_list[experiments[date_time]['condition']] = [exp_ID]

    
    return ID_dict, drugs_list


 
    
def find_surfaces(date_time, condition, peak_type, peak_criteria = 'frequency', dist = 1.5, return_channels = False) :
    '''
    TO DO 
    '''
    # we get the list of channels and the 
    channels, z = get_channels_and_z(date_time, condition, peak_criteria, peak_type)

    indices = [i for i in range (len(z)) if z[i] > min_val[peak_type][peak_criteria]]
    surf = dist*len(indices)

    if return_channels :
        channels_sup = [channels[i] for i in indices]
        return surf, channels_sup

    return surf


class compute_results :
    def __init__ (self, peak_type, conditions, list_IDs, ID_dict, compute_surface = False, N_best = spe.N_best) :
        channels_ord = read_dict(f'{spe.results}{peak_type}_chan_ord.pkl')
        self.infos = {}
        self.frequencies = {}
        self.amplitudes = {}
        self.powers = {}
        self.surfaces = {}
        self.bad_IDs = []

        for exp_ID in list_IDs : 
            date_time = ID_dict[exp_ID]['date_times'][0]
            channels = channels_ord[exp_ID][:spe.N_best]
            freq = []
            for channel in channels: 
                try :
                    freq.append(peak_info(date_time, 'baseline', channel, peak_type, 'frequency'))
                except : 
                    print(f'{date_time} :  {channel} not found')
            
            if np.nanmedian(freq) < min_val[peak_type]['frequency'] : 
                self.bad_IDs.append(exp_ID)

        for c,condition in enumerate(conditions) : 
            self.infos[condition] = []
            self.frequencies[condition] = []
            self.amplitudes[condition] = []
            self.powers[condition] = []
            self.surfaces[condition] = []
            for exp_ID in list_IDs : 
                if exp_ID not in self.bad_IDs : 
                    date_time = ID_dict[exp_ID]['date_times'][c]
                    channels = channels_ord[exp_ID][:N_best]
                    freq = []
                    ampli = []
                    power = []
                    for channel in channels: 
                        try : 
                            freq.append(peak_info(date_time, condition, channel, peak_type, 'frequency'))
                            ampli.append(peak_info(date_time, condition, channel, peak_type, 'amplitude'))
                            power.append(peak_info(date_time, condition, channel, peak_type, 'power'))
                        except : 
                            print(f'{date_time} :  {channel} not found')

                    self.infos[condition].append({'exp_ID' : exp_ID, 'date_time' : date_time, 'channels' : channels})
                    self.frequencies[condition].append(np.nanmedian(freq))
                    self.amplitudes[condition].append(np.nanmedian(ampli))
                    self.powers[condition].append(np.nanmedian(power))
                    if compute_surface :
                        self.surfaces[condition].append(find_surfaces(date_time, condition, peak_type))






def run_results(drugs, peak_types, washout = True, N_best = spe.N_best) :
    '''
    This is a function very specific to experiments with baseline / drug / washout 
    It will find for each peak type, for each drug, the main results (frequency of events, their amplitude, their power and the surface of the tissue with peaks)

    The real computation of results happens in the function peak_info, here we just use compute results to call peak_info for the n best channels 
    '''
        

    ID_dict, ID_drug = experiments_dict() # first we get the information about each experiment and the list of experiments for each drug 
    
    for peak_type in peak_types : 
        for drug in drugs : 
            print(f'--------------------- {peak_type} -- {drug} ---------------------')
            if washout :
                conditions = ['baseline', drug, 'washout']
            else : 
                conditions =  ['baseline', drug]

            exp_IDs = ID_drug[drug]
            results = compute_results(peak_type, conditions, exp_IDs, ID_dict, compute_surface = True, N_best = N_best)
            save_dict(results, f'{spe.results}{peak_type}_{drug}.pkl')

    
def check_recordings(peak_type, save = False, to_return = True) : 
    '''
    TO DO 
    '''
    gui_recordings = {}
    ID_dict, ID_drug = experiments_dict() # first we get the information about each experiment and the list of experiments for each drug
    gui_recordings = {}
    channels_ord = read_dict(f'{spe.results}{peak_type}_chan_ord.pkl')
    for drug in spe.drugs : 
        exp_IDs = ID_drug[drug]
        for exp_ID in exp_IDs : 
            channel = channels_ord[exp_ID][0]
            best_channels = channels_ord[exp_ID][:11]
            c = 0
            date_time = ID_dict[exp_ID]['date_times'][c]
            gui_recordings[date_time] = {'label' : date_time, 'value' : 0, 'var' : 0, 'channel' : channel, 'channels' : best_channels, 'exp_ID' : exp_ID}

    gui_recordings = make_choices(gui_recordings, text_keys=[], checkbox_keys=list(gui_recordings.keys()))

    date_times = list(gui_recordings.keys())
    bad_recordings = []

    for date_time in date_times  :
        if gui_recordings[date_time]['value'] == 1: 

            channel = gui_recordings[date_time]['channel']
            raw_channel = read_dict(f'{spe.results}raw_{date_time}_{channel}.pkl')
            peaks = get_peaks_from_pkl(channel, date_time)
            infos = {key : gui_recordings[date_time][key] for key in list(gui_recordings[date_time].keys()) if key not in ['value', 'var']}
            infos['peak_type'] = peak_type
            infos['n_frames'] = len(peaks.frame_index[peak_type])
            print(infos)
            data = data_time_window(raw_channel, frames_to_show = peaks.frame_index[peak_type], ylim = [-100,100], window = 15*freqs, peak_type = peak_type)
            gui_recordings['bad'] = {'label' : date_time, 'value' : '', 'var' : 0}
            gui_recordings = make_choices(gui_recordings, text_keys=['bad'], checkbox_keys=[])
            if gui_recordings['bad']['value'] != '' :
                infos['comment'] = gui_recordings['bad']['value']
                bad_recordings.append(infos)

    for record in bad_recordings : 
        print(record)
    
    if save :
        save_dict(bad_recordings, f'{peak_type}_bad_recordings.pkl')
    
    if to_return : 
        return bad_recordings



def plot_main_results(drugs, plots = {'interictal' : ['frequency', 'amplitude'], 'MUA' : ['frequency']}, peak_types = peak_types, echo = False, washout = True, highlight = False, normalize_to_baseline = False, saving = False, name_fig = 'final') :
    '''
    This function calls the box plotting function and makes plots for the different drugs with the frequency and the amplitude of each peak type based on the results obtained in 'run_results'
    Add labels and clearer option for modif plot parameters and more compact subplots
    '''

    for peak_type in peak_types : 
        fig, axes = start_fig(len(plots[peak_type]), len(drugs))

        for d,drug in enumerate(drugs) : 
            if len(drugs) == 1 : 
                ax = axes
            else : 
                ax = axes[d]

            if len(plots[peak_type]) == 1 : 
                ax = [ax]

            if washout :
                conditions = ['baseline', drug, 'washout']
            else : 
                conditions = ['baseline', drug]

            results = read_dict(f'{spe.results}{peak_type}_{drug}.pkl')
            if echo :
                for condition in conditions :
                    for info in results.infos[condition] : 
                        print(f'{peak_type} / {drug} / {condition} : {info}')
            
            to_highlight = []
            if highlight : 
                for i in range(len(results.infos['baseline'])) :
                    if results.infos['baseline'][i]['exp_ID'] in spe.to_highlight : 
                        to_highlight.append(i)

            plot = 0 
            if 'frequency' in plots[peak_type] : 
                box_plotting([results.frequencies[c] for c in conditions], conditions, peak_type, f'frequency_{peak_type}_{drug}', subplots = True, fig = fig, ax =ax[plot], to_highlight=to_highlight, normalize_to_baseline=normalize_to_baseline)
                plot += 1

            if 'amplitude' in plots[peak_type] : 
                box_plotting([results.amplitudes[c] for c in conditions], conditions, peak_type, f'amplitude_{peak_type}_{drug}', subplots = True, fig = fig, ax =ax[plot], to_highlight=to_highlight, normalize_to_baseline=normalize_to_baseline)
                plot += 1
            
            if 'surface' in plots[peak_type] : 
                box_plotting([results.surfaces[c] for c in conditions], conditions, peak_type, f'surface_{peak_type}_{drug}', subplots = True, fig = fig, ax =ax[plot], to_highlight=to_highlight, normalize_to_baseline=normalize_to_baseline)
                plot += 1

            if 'power' in plots[peak_type] : 
                box_plotting([results.powers[c] for c in conditions], conditions, peak_type, f'surface_{peak_type}_{drug}', subplots = True, fig = fig, ax =ax[plot], to_highlight=to_highlight, normalize_to_baseline=normalize_to_baseline)
                plot += 1
            
        if saving :
            plt.savefig(f'{spe.figures}{name_fig}_{peak_type}.{format_fig}', format = format_fig)


def statistics(y, show_ns = False, print_stats = False) :
    '''
    From a list of results y (with each element of y are the same size), returns statistics 
    Only works on paired test for now, need to update (todo)
    Input: y list of lists of same len to statistically compare
          test a function for comparing the different groups 
    Ouput: stars list of significant differences for each [index of reference group in y, index of compared group in y, stars to put depending on significance level]
    '''
    def test_before_ttest(A,B, pval = 0.05) :
        sampling_difference = np.array(A) - np.array(B)
        _,p1 = stats.shapiro(sampling_difference)
        _, p2 = stats.levene(A,B,center= 'mean')
        if p1 > pval and p2 > pval :
            return True
        else :
            return False

    stars = [] 

    # Go through lists to compare 
    for i in range (len(y)) :
        # Compare one list to all the next ones 
        for j in range (i+1, len(y)) : 
            emptytest1 = [k for k in y[i] if k not in [None, 0, np.nan]]
            emptytest2 = [k for k in y[j] if k not in [None, 0, np.nan]]
            if len(emptytest1) > 2 and len(emptytest2) > 2 : 
                if test_before_ttest(y[i],y[j]) :
                    test = 'ttest'
                    pval = stats.ttest_ind(y[i], y[j]).pvalue # Find the pvalue
                else :
                    test = 'wilcoxon'
                    pval = stats.wilcoxon(y[i], y[j]).pvalue # Find the pvalue

                if print_stats : 
                    print(i,j,pval, test)
                # Assign number of stars depending on the pvalue 
                if pval < 0.001 :
                    stars.append([i,j,'***', pval, test]) 
                elif pval < 0.01 :
                    stars.append([i,j,'**', pval, test])
                elif pval < 0.05 :
                    stars.append([i,j,'*', pval, test])
                elif show_ns : 
                    stars.append([i,j,'ns', pval, test])
                
    return stars 


## PLOTTING 

    
def box_plotting(y, conditions, peak_type, name = '', normalize_to_baseline = False, saving = False, subplots = False, fig = None, ax = None, to_highlight = []) : 
    '''
    This function is used to make all the boxplots from ys
    ys : (baseline, drug, washout) data
    check: dict with {'criteria': name of the criteria, 'category' : where to find this condition in ID_dict, 'condition': the experimental condition to check}
            multiple_channels: True/False depending on if you want to take only the best channels or the n_best channels stored in ID_dict

    todo: commment more and adapt for time_limit and plot in time do not work anymore, need to put them in other function see pipeline report task3
    '''

    print(f'{peak_type} - {name} - {len(y[0])} trials')


    if subplots and ax is None :
        fig, ax = start_fig()

    try :
        box = np.array(y)
        
    except :
        print('error in size of y')
        print(y)
        box = y
    
    box = [yi[~np.isnan(yi)].tolist() for yi in box]
    

    stars = statistics(box, show_ns = False, print_stats = True)

    if normalize_to_baseline :
        ref = np.nanmedian(box[0])
        if np.isnan(ref) or ref == 0 :
            ref = 1
            print('!!!NO NORMALIZATION!!!!')
        box = [[boxi[i]/ref for i in range (len(boxi))] for boxi in box]
    
    if subplots :
        ax = make_boxplot(box, conditions, draw_line=True, name_fig=name, stars = stars, saving = saving, fig = fig, ax = ax, show = False, to_highlight=to_highlight)

    else :
        make_boxplot(box, conditions, draw_line=True, name_fig=name, stars = stars, saving = saving, to_highlight=to_highlight)

    if subplots :
        return ax
    else : 
        plt.show()



# Plot signal / peaks 


def select_peaks(path: string, channel:string ,peak_types: list = peak_types, x_window:dict = {'interictal' : 10000, 'MUA' : 500}, peaks = None) -> list: 
    '''
    This function creates a plot with the raw signal, the filtered signal and the peaks found 

    Input: 
    path of the recording
    channel of interest 
    x_window (ms): the zoom we want when we look at the peaks for each peak type (for ex, default value for interictal is 10000 so we see 10s) 


    TO do use the threshold in peaks.thresholds[peak_type] instead of finding a new one 
    Output:
    deleted: indices of the deleted peaks  

    '''

    date_time = date_time_from_path(path)

    if peaks is None :
        peaks = read_dict(f'{spe.peaks}{date_time}_{channel}.pkl')
    

    

    print(f'{date_time}_{channel} processing ...')

    # Get the raw data
    raw, channels = raw_signals_from_file(path)
    channel_index = channels.index(channel)

    # Take only the signal from the channel chosen and convert to µV
    raw_channel = factor_amp*raw[channel_index]

    # Call the class preprocessing to filter the signal 
    data = preprocessing(raw_channel, peak_types)
    

    # We put the signals of interest in the signals variable for each data type and peak_types 
    signals = {}
    for data_type in data_types : 
        # Get the filtered and normalised signals for each peak type
        if data_type != 'raw':
                signals[data_type] = data.signal[data_type]
        else : # For the raw data we have the same signal for all peak types 
            signals[data_type] = {}
            for peak_type in peak_types:
                signals[data_type][peak_type] = data.raw

    

    frame_index = peaks.frame_index
    
    amplitude = peaks.amplitude

    deleted = {}
    
    # For each peak type

    for peak_type in peak_types:

        if len(frame_index[peak_type]) > 0 : 

            for data_type in data_types :
                amplitude[peak_type][data_type] = [amp for amp in amplitude[peak_type][data_type] if not isnan(amp)]
                if len(frame_index[peak_type]) != len(amplitude[peak_type][data_type]) :
                    print('ERROR in peak file with number of peaks')
                

            print(f'Plotting the signal with {peak_type} peaks...')
            # Create the figure 
            fig, ax = start_fig(nrows=len(data_types))

            # Plot parameters 
            signal_ref = signals[data_type][peak_type]
            x_ms = [i/10 for i in range (len(signal_ref))]
            param_plot = {
                'xlim' : [x_ms[0], x_ms[-1]],
                'xlabel' : 'Time',
                'ylim' : [[int(min(signals[data_type][peak_type])), int(max(signals[data_type][peak_type]))] for data_type in signals.keys()],
                'ylabel' : 'Amplitude',
                'x_window': x_window[peak_type],
                'ticks' : [],
                'y_zoom': [0.25,5]
                }

            # We add the list of peaks to the plot parameters
            if len(frame_index[peak_type]) > 0 :   
                ms_index = [x_ms[i] for i in frame_index[peak_type]]
                param_plot['ticks'] = list(ms_index)

            for s,data_type in enumerate(signals.keys()) :
                ax[s].set_title(f'{data_type} signal')
                # Plot the signal with the right color for each peak type 
                ax[s].plot(x_ms, signals[data_type][peak_type], color = param[peak_type]['color_plot'], alpha = alpha_plots)

                # If peaks were found, show them as stars with the right color for each peak type 
                if len(frame_index[peak_type]) > 0 :   
                    ax[s].scatter(ms_index, np.array(amplitude[peak_type][data_type]), marker = '*', color = param[peak_type]['color_peak'])
                
                if data_type == 'absolute' : 
                    # Show an horizontal line for threshold with the right color for each peak type 
                    ax[s].plot([data.threshold[peak_type] for _ in range(len(signals[data_type][peak_type]))], color = param[peak_type]['color_peak'])
                    
                # Definition of the x and y axes limits 
                ax[s].set_xlim(param_plot['xlim'][0], param_plot['xlim'][0]+param_plot['x_window'])
                ax[s].set_ylim(param_plot['ylim'][s])

            # Create sliding bars to go through different time points, zoom in and out in amplitude and time 
            n_peaks = len(frame_index[peak_type])
            deleted[peak_type] = make_it_interactive(fig, ax, param_plot)
        
            # Print number of peaks deleted
            n_del = len(deleted[peak_type])
            print(f'{n_del} peaks out of {n_peaks} total peaks have been deleted') 

    return peaks, deleted 

def why_not_this_channel(date_time, channel, peak_type) : 
    '''
    TO DO 
    '''
    peaks = read_dict(f'{spe.peaks}{date_time}_{channel}.pkl')
    frames = peaks.frame_index[peak_type]
    try : 
        raw_channel = read_dict(f'{spe.results}raw_{date_time}_{channel}.pkl')
    except : 
        start, stop = determine_timing(date_time, condition='baseline')
        path = path_from_datetime(date_time)
        save_raw_data(path, channel, [start,stop])
        raw_channel = read_dict(f'{spe.results}raw_{date_time}_{channel}.pkl')
        
    data = data_time_window(raw_channel, plot_filtered_raw= True, plot_raw= False, frames_to_show=frames, peak_type=peak_type)

class plot_signal : 
    def __init__ (self, path: string, channel:string, frames = [0,10*freqs], name = None, scale = {'x' : None, 'y' : None}) :
        self.path = path
        self.date_time = date_time_from_path(path)
        self.channel = channel
        self.frames = frames
        self.scale = scale 
        raw = self.get_raw_channel()
        self.data = preprocessing(raw, peak_types)

        if name is None : 
            self.name = f'{self.date_time}_{self.channel}_signal_plot'

    def create_plot(self, raster_plot, data_types = ['raw'], type_plot = '') : 
        if raster_plot :
            height_ratios = [3 for _ in range (len(data_types)*len(peak_types) -1)] + [1 for _ in range (len(peak_types))] + [3]
            fig, axes = start_fig(nrows = 2+len(peak_types), height_ratios = height_ratios, figsize = (3*(cm2inch(fig_len)), 4 * (cm2inch(fig_len))))
        else : 
            fig, axes = start_fig(nrows = 2, figsize = (4*(cm2inch(fig_len)), 2* (cm2inch(fig_len))))

        fig.tight_layout(pad=0)

        for ax in axes :
            ax = set_ax_parameters(ax, just_plot=True)

        try :
            param_file = read_dict(f'{spe.param}//{self.name}{type_plot}.pkl')
        except :
            param_file = initialize_plot_parameters(f'{self.name}') 

        self.plot_param = {}
        for key in param_file.keys() : 
            if param_file[key]['value'] == '' :
                self.plot_param[key] = None
            else : 
                self.plot_param[key] = param_file[key]['value']

        return fig, axes 
    

    def adjust_axes(self, axes) : 

        if self.plot_param['xlim'] is None : 
            self.xlim = axes[0].get_xlim()
        else : 
            self.xlim = self.plot_param['xlim']
        
        if self.plot_param['ylim'] is None : 
            self.ylim = axes[0].get_ylim()
        else : 
            self.ylim = self.plot_param['ylim']

        for ax in axes : 
            ax.set_xlim(self.xlim)
            ax.set_ylim(self.ylim)
        
        return axes


    def add_raster_plot(self, axes) : 
        frame_ind, _, _ = get_peaks_from_pkl(self.channel, self.date_time)
        for p,peak_type in enumerate(peak_types) : 
            frame_index = [f-self.frames[0] for f in frame_ind[peak_type] if f in range (self.frames[0], self.frames[1])]
            for frame in frame_index : 
                axes[-2-p].plot([frame,frame], [0,1], color = 'black', lw = 2)
            axes[-2-p].set_ylabel(peak_type)
        
        return axes

    def add_scale(self, axes) : 
        if self.scale['x'] is None :
            xbar = (self.frames[1]- self.frames[0])/10
            print(f'Ref in x: {xbar/freqs} s')
        if self.scale['x'] is None :
            ybar = (self.ylim[1]-self.y_lim[0])/10
            print(f'Ref in y: {ybar} µV')

        axes[-1].plot([self.xlim[0], self.xlim[0]+xbar], [0,0], color = 'black')
        axes[-1].plot([self.xlim[0]+xbar, self.xlim[0]+xbar], [0,ybar], color = 'black')

        return axes


    def get_raw_channel(self) : 
        try :
            raw_channel = f'{spe.results}raw_{self.date_time}_{self.channel}.pkl'
        
        except : 
            if self.date_time in spe.specials.keys() :
                self.path = spe.specials[self.date_time]['path']

            raw, channels = raw_signals_from_file(self.path)
            channel_index = channels.index(self.channel)
            raw_channel = factor_amp*raw[channel_index]

        raw_channel = raw_channel[self.frames[0]:self.frames[1]]
        return raw_channel



    def plot_the_plot(self, data_types = ['raw'], peak_types = peak_types, saving = False, format_fig = format_fig, raster_plot = False) : 

        _, axes = self.create_plot(raster_plot = raster_plot, data_types = ['raw'], type_plot = '_raw')
        
        n = 0
        for data_type in data_types :
            if data_type == 'raw' :
                axes[n].plot(self.data.raw, color = 'black')
                n += 1
            else : 
                for peak_type in peak_types : 
                    axes[n].plot(self.data.signal[data_type][peak_type], color = param[peak_type]['color_plot'])
                    n += 1                
        
        axes = self.adjust_axes(axes)

        if raster_plot :
            axes = self.add_raster_plot(axes)

        axes = self.add_scale(axes)

        if saving : 
            plt.savefig(f'{spe.figures}{self.name}_raw.{format_fig}', format = format_fig, transparent = True)
        plt.show()

    def time_spectrogram(self) :
        window = signal.windows.kaiser(256, beta=5)
        nperseg = 256
        noverlap = 128
        nfft = 512
        f, t, Sxx = signal.spectrogram(self.data.raw, fs = freqs, window=window, nperseg=nperseg, noverlap=noverlap, nfft=nfft)
        fig, ax = start_fig()
        image = ax.pcolormesh(t, f, Sxx, shading='gouraud', cmap= 'inferno')
        ax.plot(self.data.raw, color = 'black', lw = 0.5)
        fig.colorbar(image, ax=ax, location = 'top', fraction=0.6, pad=0)
        fig.tight_layout(pad=0)




def subplot_all_channels(path, frames = [0,freqs], ylim = None, layout = 'sparse', letter_lim = None, number_lim = None, saving = False, name_fig = 'all_channels') :
    '''
    This function shows the raw data for all the channels as a big matrix 
    '''
    date_time = date_time_from_path(path)
    try :
        raw_channels = f'{spe.figures}{date_time}_raw_channels.pkl'
    except :
        raw_channels = raw_data_all_channels(path, frames, to_return = True)

    alphabet = list(string.ascii_uppercase)
    i = alphabet.index('I')
    del alphabet[i]
    channels = ', '.join(list_channels_all[layout])
    
    if letter_lim is None : 
        letters = re.findall(r'[a-zA-Z]',channels)
        min_letter = alphabet.index(min(letters))
        max_letter = alphabet.index(max(letters))
    else : 
        min_letter = alphabet.index(letter_lim[0])
        max_letter = alphabet.index(letter_lim[1])

    rangeletter = max_letter - min_letter + 1

    if number_lim is None : 
        numbers = re.findall(r'\d+',channels)
        numbers = [int(number) for number in numbers]
    else : 
        numbers = number_lim
    
    rangenumber = max(numbers) - min(numbers) + 1 

    fig, axes = start_fig(nrows = rangenumber, ncols = rangeletter)
    
    fig.tight_layout()

    for n in range (min(numbers), max(numbers) + 1) : 
        for letter in range (min_letter, max_letter + 1) :
            raw_channel = raw_channels[n][alphabet[letter]]
            axes[n-min(numbers)][letter-min_letter].plot(raw_channel, color = 'black')
            if ylim is not None :
                axes[n-min(numbers)][letter-min_letter].set_ylim(ylim)
            axes[n-min(numbers)][letter-min_letter].set_yticks([],[])
            axes[n-min(numbers)][letter-min_letter].set_xticks([],[])
        print(f'{(n+1)*100/rangenumber}% done', end = '\r')

    if saving :
        plt.savefig(name_fig, format = format_fig)
    plt.show()


# Quantitative plots 

def show_colormap(date_time: string, condition: string, peak_criteria: string, peak_types, saving:bool = False, format_save: str = format_fig, show_channels = False, echo = False) : 
    '''
    date_time date and time of the recording of interest 
    peak_criteria: one of the following ['frequency', 'amplitude', 'power']

    update to do : make 1z and 2z as one function 
    No output 
    TO DO add plot_parameters
    '''

    # Find the path of the image from the date_time of the recording 
    image = get_image_from_info(date_time)
    
    #  Get the positions of the channels 
    channels, x,y = get_channels_positions(date_time)

    # Find out the letters and numbers present in the channels names 
    alphabet = list(string.ascii_uppercase)
    letters = [letter for letter in alphabet if letter in [channel[0] for channel in channels]]
    numbers = range(1,max([int(channel[1:]) for channel in channels])+1)

    # Initialize the variables 
    X = [[] for _ in range (len(letters))]
    Y = [[] for _ in range (len(numbers))]# We create a grid with the different channel numbers as rows (each row has a different y value in Y)
    Z = [[0]*len(numbers) for _ in range (len(peak_types))] # Each channel in the grid has a z value based on the peaks frequency or the duration of the bursts recorded in this channel... 
    not_analyzed = [[] for _ in range (len(peak_types))]
    channels_analysed = [[] for _ in range (len(peak_types))]
    z = [[] for _ in range (len(peak_types))]
    min_z = [0 for _ in range(len(peak_types))]
    max_z = [0 for _ in range(len(peak_types))]

    # Based on the decided criteria, get the right value for each channel 
    for p,peak_type in enumerate(peak_types) : 
        channels_analysed[p], z[p] = get_channels_and_z(date_time, condition, peak_criteria, peak_type)
        min_z[p] = np.nanmin(z[p])
        max_z[p] = np.nanmax(z[p])
        for number in numbers : 
            Z[p][number-1] = [0]*len(letters) 
            for l,letter in enumerate(letters): 
            
                channel = f'{letter}{number}' 
                if channel in channels: 
                    ind = channels.index(channel) # We find the index of the channel of interest in the original channel list (not ordered)

                    X[l].append(x[ind])
                    Y[number-1].append(y[ind])

                    if channel in channels_analysed[p]  : 
                        ind_analysed = channels_analysed[p].index(channel)
                        Z[p][number-1][l] = z[p][ind_analysed] # We keep the z value of this channel in the right place 
                        if z[p][ind_analysed] > min_val[peak_type][peak_criteria] and echo :
                            print(f'{channel} : ({x[ind]}, {y[ind]}), z = {z[p][ind_analysed]}')
                    else :
                        not_analyzed[p].append(channel)

        if echo :
            print(f'{peak_type} : Channels {not_analyzed[p]} were not analysed')

    Y= np.nanmedian(Y, axis = 1) # mean y coordinate through all letters for this number 
    X = np.nanmedian(X, axis = 1) # mean x coordinate through all numbers for each letter


    # Open the image 
    I = np.asarray(Image.open(image))
    
    # We plot the image and put a transparent (based on alpha value) colormap above it depending on the z values
    _, ax = start_fig(figsize = (3 * (cm2inch(fig_len)), 2 * (cm2inch(fig_len))))
    xlim = 0,I.shape[0]
    ylim = 0,I.shape[1]

    if show_channels : 
        for number in numbers: # For each channel number 
            for l,letter in enumerate(letters) : 
                ax.text(X[l], Y[number-1], f'{letter}{number}', color = 'black')
    
    ax.imshow(I)
    ax = set_ax_parameters(ax, just_plot=True)
    divider = make_axes_locatable(ax)

    mapcolors = [None for _ in range (len(peak_types))]
    for_int = 10000
    steps = [None for _ in range (len(peak_types))]
    levels = [None for _ in range (len(peak_types))]
    cbs = [None for _ in range (len(peak_types))]
    col_ax = [None for _ in range (len(peak_types))]
    cbars = [None for _ in range (len(peak_types))]
    for p,peak_type in enumerate(peak_types) : 
        Yg, Xg, Z[p] = expand_matrix(list(Y), ylim, list(X), xlim, Z[p]) # put values everywhere for homogeneity (NB: here Z is defined with (Y,X) not (X,Y) be careful)
        mapcolors[p] = colormaps[peak_type]
        mapcolors[p].set_under(color='white', alpha = 1)
        mapcolors[p] = add_linear_transparency(mapcolors[p])
        steps[p] = int((for_int*max_z[p] - for_int*min_z[p])/steps_color)
        levels[p] = np.array([i/for_int for i in range (int(for_int*min_z[p]), int(for_int*max_z[p]), steps[p])])
        cbs[p] = ax.contourf(Xg, Yg, Z[p], cmap= mapcolors[p], levels = levels[p], antialiased=True, vmin = min_z[p], vmax =max_z[p],  algorithm = 'serial')

        col_ax[p] = divider.append_axes(ax_cmap[peak_type]['side'], size="5%", pad=0)
        steptick = (int(for_int*max_z[p]) - int(for_int*min_z[p]))/5
        colorticks= [round(i/for_int,3) for i in range (int(for_int*min_z[p]), int(for_int*max_z[p])+int(steptick), int(steptick))]
        cbars[p] = plt.colorbar(cbs[p], cax=col_ax[p], ticks = colorticks)
        cbars[p].ax.tick_params(labelsize= fontsize)
        cbars[p].set_label(f'{labels[peak_type]}  {labels[peak_criteria]}', rotation=ax_cmap[peak_type]['rotation'], fontsize = fontsize, labelpad = 20, font = font)
        cbars[p].ax.set_yticklabels(colorticks)

    if saving :
        plt.savefig(f'{spe.figures}{date_time}_{peak_criteria}.{format_save}', format = format_save)

    plt.show()



def make_boxplot(y, experiments_keys, name_fig = 'plot', saving = False, draw_line = False, stars = [], fig = None, ax = None, show = True, colors_plot = None, to_highlight = []) : 
    '''
    Creates boxplot based on values in y and a list of keys corresponding (y and experiments_keys have the same len)
    y : list of lists each list contains the values for each condition 
    experiments_keys : list of names of the conditions (with associated colors in colors_plot in the parameters_file)
    name_fig: the name of the figure for saving 
    Put saving = True so the figure is saved 

    todo adjust the height of the stat bar so no overlap with dots 
    to do improbe to highlight so works with broken axis 

    ADD some comments 
    '''

    for i in range (len(y)) :
        print(f'{name_fig} - {experiments_keys[i]}. Median : {np.nanmedian(y[i])}. Std : {np.nanstd(y[i])}, Mean: {np.nanmean(y[i])}, SEM: {sem(y[i])}')

    for star in stars :
        print(f'{experiments_keys[star[0]]} and {experiments_keys[star[1]]}, p = {star[3]}   ({star[4]})')
    
    # Get the right colors 
    if colors_plot is None : 
        try :
            colors_plot = [spe.colors[exp] for exp in experiments_keys]
        except :
            colors_plot = ['white' for _ in experiments_keys]
    median_colors = [spe.inverse_color[color] for color in colors_plot]

    try :
        param_file = read_dict(f'{spe.param}//{name_fig}.pkl')
    except :
        param_file = initialize_plot_parameters(name_fig) 

    plot_param = {}
    for key in param_file.keys() : 
        if param_file[key]['value'] == '' :
            plot_param[key] = None
        else : 
            plot_param[key] = param_file[key]['value']


    if ax is None :
        fig, ax = start_fig(1, 1)
    
    
    bp = ax.boxplot(y, whis = whis, widths=width_bp, showfliers=False, patch_artist=True)
    
    ylim = plot_param['ylim']

    if ylim is None :
        ylim = ax.get_ylim()
        maxi = np.nanmax([np.nanmax(yi) for yi in y])
        mini = np.nanmin([np.nanmin(yi) for yi in y])
        if not (np.isnan(mini)) and not (np.isinf(mini)) : 
            ylim = [min(mini,ylim[0])*0.9, max(ylim[1], maxi)*1.1]


    ax.set_ylim(ylim)
    
    x = [[i + 1.1 + width_bp/2 for _ in range(len(y[i]))] for i in range (len(experiments_keys))]
    x = [space_dot_boxplot(xi,yi, ylim) for xi,yi in zip(x,y)]

    if draw_line: 
        for j in range (len(y[0])) :
            xline = []
            yline = []
            for i in range (len(experiments_keys)) :
                xline.append(x[i][j])
                yline.append(y[i][j])
            ax.plot(xline,yline, color = 'grey', alpha = alpha_plots)
    

    ax = set_ax_parameters(ax, yticks = plot_param['yticks'], ylabel = plot_param['ylabel'], xticks = plot_param['xticks'])
    
    points = []
    for xi, yi,c in zip(x,y, colors_plot) :
        points.append(ax.scatter(xi, yi, color = c, edgecolors=edgecolor, s = cm2pts(dot_size)))
    
    for p in to_highlight :  
        for i in range (len(x)) :
            ax.scatter(x[i][p], y[i][p], color = 'red', edgecolors=edgecolor, s = cm2pts(dot_size))


    for patch, color, median, median_color in zip(bp['boxes'], colors_plot, bp['medians'], median_colors):
        patch.set_facecolor(color)
        patch.set_linewidth(lw)
        median.set_color(median_color)
        median.set_linewidth(lw_median)

    if stars is not None : 
        ax = add_stats(ax, stars, ylim)
    
    broken = plot_param['broken_y_axis']
    if broken is not None :
        add_broken_axis(fig, ax, [ylim, broken], points = points, box = bp, boxcolors = colors_plot, stars = stars)
    


    if saving :
        name_fig = name_fig.replace(' ','_')
        plt.savefig(f'{spe.figures}{name_fig}.{format_fig}', format = format_fig, transparent = True)
    
    print(f'{name_fig} - y ticks : {ax.get_yticks()}')

    if show : 
        plt.show()
    else : 
        return ax 



## Useful for plots 


def cm2inch(x):
    return x/2.54

def cm2pts(x) : 
    return x*28.35

def pts2inch(x) : 
    return x/72

def initialize_plot_parameters(name_plot) :
    '''
    TO DO 
    '''
    save_dict(gui_plot, f'{spe.param}//{name_plot}.pkl')
    return gui_plot

def change_plot_parameters(name_plot) : 
    '''
    TO DO 
    '''
    plot_param = read_dict(f'{spe.param}//{name_plot}.pkl')
    for key in plot_param.keys() :
        if type(plot_param[key]['value']) == list :
            plot_param[key]['value'] = ', '.join([str(i) for i in plot_param[key]['value']])
    to_modify = parameters_comma + parameters_text
    plot_param = make_choices(plot_param, text_keys = to_modify, checkbox_keys = [])
    for modify in parameters_comma :
        plot_param[modify]['var'] = []
        if plot_param[modify]['value'] != '' : 
            try :
                plot_param[modify]['value'] = [float(idx) for idx in plot_param[modify]['value'].replace(' ', '').split(',')]
            except : 
                plot_param[modify]['value'] = [idx for idx in plot_param[modify]['value'].replace(' ', '').split(',')]
    for modify in parameters_text :
        plot_param[modify]['value'] = plot_param[modify]['value']
        plot_param[modify]['var'] = []

    save_dict(plot_param, f'{spe.param}//{name_plot}.pkl')

    

def start_fig(ncols = 1, nrows = 1, height_ratios = None, width_ratios  = None, pad = 0.8, figsize = None) :
    if figsize is None :
        figsize = (ncols * (cm2inch(fig_len)+pts2inch(3*pad)), nrows * (cm2inch(fig_len) + pts2inch(pad)))
    f, ax = plt.subplots(ncols = ncols, nrows= nrows, figsize = figsize, gridspec_kw={'height_ratios':height_ratios, 'width_ratios' : width_ratios})
    f.tight_layout(pad=pad, w_pad=3*pad, h_pad=pad)
    return f, ax 

def set_ax_parameters(ax, nticks = None, yticks = None, just_plot = False, xlabel = None, ylabel = None, xticks = None, show_xticks = False, show_yticks = True) :
    '''
    For all my plots, I need to delete the top and right lines on the plot, use a certain tick length and width, and no label on the ticks (but need to see their value to put them in Illustrator later)
    Input: ax to change
    Output: changed ax 
    '''

    ax.spines[['top','right']].set_visible(False) # remove the black lines 

    if xticks is not None :
        show_xticks = True
        if type(xticks[0]) == str :
            xticks_pos = [i for i in range(1,len(xticks)+1)]
        else : 
            xticks_pos = xticks 

    if show_xticks :
        if xticks is None : 
            xticks = ax.get_xticks()
            if nticks is not None : 
                xstep = (max(xticks)-min(xticks))/nticks
                xticks = [min(xticks) + i*xstep for i in range (nticks)]
                print(xticks)
                xticks = rounding(xticks)
                print(xticks)
                ax.set_xticks(xticks[1:], xticks[1:], font = font, fontsize = fontsize)
            else : 
                ax.set_xticks(xticks, xticks, font = font, fontsize = fontsize)
        else  :
            ax.set_xticks(xticks_pos, xticks, font = font, fontsize = fontsize)

        
        ax.tick_params(axis="x", length = 0)
        #ax.tick_params(axis="x",direction="in", length = cm2pts(tick_len), width = tick_thick) # ticks going in the plot at the right size 

    else : 
        ax.set_xticks([]) # For the x axis no ticks at all


    if show_yticks :
        if yticks is None :
            yticks = ax.get_yticks()
            if nticks is not None : 
                ystep = (max(yticks)-min(yticks))/nticks
                yticks = [min(yticks) + i*ystep for i in range (nticks)]
                yticks = rounding(yticks)
                yticks = yticks[1:]
        yticks = [int(yi) if (yi%1 == 0) else yi for yi in yticks]
        ax.set_yticks(yticks, yticks, font = font, fontsize = fontsize)
        ax.tick_params(axis="y",direction="in", length = cm2pts(tick_len), width = tick_thick) # ticks going in the plot at the right size 
    else :
        ax.set_yticks([])

    if ylabel is not None :
        ax.set_ylabel(ylabel, font = font, fontsize = fontsize)

    if xlabel is not None :
        ax.set_xlabel(xlabel, font = font, fontsize = fontsize)
    
    if just_plot : 
        ax.set_xticks([])
        ax.set_yticks([])
        ax.spines[['bottom', 'left']].set_visible(False) # remove the black lines 
    
    return (ax)


def space_dot_boxplot(x,y,ylim) :
    '''
    Messy function to put some horizontal space between dots that are too close together in scattering

    Input: x = list of coordinates to space if need 
           y = list of fixed coordinates
           ylim = [min y axis, max y axis]
    '''
    toignore = []
    try : 
        distanceval = abs(ylim[1]-ylim[0])/20
    except :
        distanceval = 0.1

    for i in range (len(y)) :
        if i not in toignore : # This is a new value we encounter 
            c = [i]
            for j in range (i+1, len(y)) : # Create a list of all values too close to the value i of interest
                if abs(y[i]-y[j]) < distanceval :
                    c.append(j)
                    toignore.append(j) # We don't need to look at these values again later on 
            
            if len(c) == 2 : # 2 values close together, one a bit on the left, one on the right 
                x[c[0]] -= 0.025
                x[c[1]] += 0.025
            if len(c) == 3 :  # 3 values close together, leaves the middle one, and put one a bit on the left, one on the right 
                x[c[0]] -= 0.05
                x[c[2]] += 0.05
            if len(c) == 4 :  # 4 values close together, 2 a bit on the left, 2 on the right (with a bit space between each)
                x[c[0]] -= 0.05
                x[c[1]] -= 0.03
                x[c[2]] += 0.02
                x[c[3]] += 0.05
            if len(c) > 4 : # More than 5 values ?! Well just put them by turn one one the left one on the right at random distances
                f = -1
                for j in range (len(c)) :
                    x[c[j]] += f*(randint(2,8))/100
                f = 1 if f == -1 else -1 
    return x 

def add_linear_transparency(cmap) :
    from matplotlib.colors import ListedColormap
    my_cmap = cmap(np.arange(cmap.N))
    my_cmap[:, -1] = np.linspace(0, 1, cmap.N)
    my_cmap = ListedColormap(my_cmap)
    return my_cmap


    
def add_broken_axis(fig, ax, limits, points = None, box = None, boxcolors = None, stars = None):
    
    # Create two subplots and hide the spines between them
    axes = fig.axes 
    index = None

    for i in range (len(axes)) :
        try : 
            for j in range(len(axes[i])) :
                if axes[i][j] == ax :
                    index = i*len(axes[i])+j
        except :
            if axes[i] == ax :
                index = i 
    
    if index is None :
        print('AX NOT FOUND')
        fig, [ax1,ax2] = start_fig(ncols = 1, nrows = 2)
    
    else :
        nrows = ax.get_gridspec().nrows
        ncols = ax.get_gridspec().ncols
        ax1 = fig.add_subplot(nrows, ncols, index+1)
        pos = ax.get_position()
        ax1.set_position([pos.x0, pos.y0+4*pos.height/7,  pos.width, 3*pos.height/7])
        ax2 = ax
        ax2.set_position([pos.x0, pos.y0,  pos.width, 3*pos.height/7])
        ax2.set_ylabel('')


    # Set the limits of each subplot
    ax2.set_ylim(limits[0][0], limits[0][1])
    ax1.set_ylim(limits[1][0], limits[1][1])


    # Plot the data on both subplots
    lines = ax.get_lines()
    for line in lines  :
        ax1.plot(line.get_xdata(), line.get_ydata(), color=line.get_color(), alpha = line.get_alpha(), lw = line.get_linewidth())
        ax2.plot(line.get_xdata(), line.get_ydata(), color=line.get_color(), alpha = line.get_alpha(), lw = line.get_linewidth())

    if points is not None :
        for point in points :
            offsets = point.get_offsets()
            x = offsets[:,0]
            y = offsets[:,1]
            colors = point.get_facecolors()
            sizes = point.get_sizes()
            ax1.scatter(x,y, c = colors, s = sizes, edgecolor = edgecolor)
            ax2.scatter(x,y, c = colors, s = sizes, edgecolor = edgecolor)
    
    if box is not None :
        medians = box['medians']
        whisks = box['whiskers']

        ydata = []
        for y in whisks:
            ydata.append(y.get_ydata())

        n = 0
        for x in medians :
            minx, maxx = x.get_xdata()
            
            miny = ydata[2*n][0]
            maxy = ydata[2*n+1][0]

            ax1.add_patch(pt.Rectangle((minx,miny), maxx-minx, maxy-miny, facecolor = boxcolors[n], edgecolor = edgecolor))
            ax2.add_patch(pt.Rectangle((minx,miny), maxx-minx, maxy-miny, facecolor = boxcolors[n], edgecolor = edgecolor))

            n += 1

    # Add slanted lines to show the break in the axis
    d = .015
    kwargs = dict(transform=ax1.transAxes, color='k', clip_on=False)
    ax1.plot((-d,+d), (-d,+d), **kwargs)
    kwargs.update(transform=ax2.transAxes)
    ax2.plot((-d,+d), (1-d,1+d), **kwargs)


    ax1 = set_ax_parameters(ax1, nticks=2, yticks = limits[1])
    ax2 = set_ax_parameters(ax2, nticks = 2, yticks = limits[0])
    
    ax2.spines['top'].set_visible(True)
    ax1.spines['bottom'].set_linestyle((0, (5, 10)))
    ax2.spines['top'].set_linestyle((0, (5, 10)))

    if stars is not None :
        ax1 = add_stats(ax1, stars, limits[1])

def add_stats(ax, stars, ylim) : 
    up = ylim[1]
    y_range = ylim[1]-ylim[0]
    for i, star in enumerate(stars):
        # Columns corresponding to the datasets of interest
        x1 = star[0] + 1
        x2 = star[1] + 1
        # What level is this bar among the bars above the plot?
        level = len(stars) - i
        # Plot the bar
        bar_height = (y_range * 0.2 * level) + ylim[1]
        bar_tips = bar_height - (y_range * 0.02)
        ax.plot(
                    [x1, x1, x2, x2],
                    [bar_tips, bar_height, bar_height, bar_tips], lw=1, c='k'
                )
        text_height = bar_height + (y_range * 0.01)
        ax.text((x1 + x2) * 0.5, text_height, star[2], ha='center', va='bottom', c='k')
        up = max(up,text_height)

    ax.set_ylim([ylim[0], up])
    return (ax)

class data_time_window() : 
    def __init__(self, data, window = 5*freqs, get_frames = False, ylim = [-100,100], frames_to_show = None, save_click = False, peak_type = 'interictal', plot_raw = True, plot_filtered_raw = False) : 
        '''
        TO DO 
        '''
        # The parametrized function to be plotted
        self.data = data
        self.window = window
        self.get_frames = get_frames 
        self.peak_type = peak_type

        if ylim is not None : 
            self.ylim = ylim 
        else : 
            self.ylim = [min(self.data), max(self.data)]

        self.frames_to_show = frames_to_show
        self.save_click = save_click

        # Define initial parameters
        self.t = 0
        self.pause = True
        plt.ion()

        if plot_raw : 
            self.fig, ax = plt.subplots()
            self.axes = [ax]
            self.plotting_raw(ax)
        
        if plot_filtered_raw :
            self.fig, self.axes = plt.subplots(2,1)
            self.plotting_raw(self.axes[0])
            data = preprocessing(self.data, peak_types = [self.peak_type])
            self.axes[1].plot(list(range(len(self.data))),data.signal['absolute'][self.peak_type], lw=2, c = 'black')
            self.axes[1].plot([data.threshold[peak_type] for _ in range(len(data.signal['absolute'][self.peak_type]))], color = 'red')
            

        # Create `matplotlib.widgets.Button` to go to previous/next time window
        axprev = self.fig.add_axes([0.3, 0.025, 0.1, 0.04])
        self.button_prev = wdg.Button(axprev, 'Previous', hovercolor='0.975') 
        self.button_prev.on_clicked(self.previous)

        axnext = self.fig.add_axes([0.8, 0.025, 0.1, 0.04])
        self.button_next = wdg.Button(axnext, 'Next', hovercolor='0.975') ## VERY important to use self.button
        self.button_next.on_clicked(self.next)
        self.fig.subplots_adjust(bottom=0.25)

        self.fig.canvas.mpl_connect('close_event', self.on_close)
        
        while self.pause : # important when we use .py files 
            plt.pause(0.1)
            

    def ax_adjust(self, ax) : 
        ax.set_xlim([self.t,self.t+self.window-1])
        ax.spines[['top','right']].set_visible(False)

    def plotting_raw(self, ax) : 

        ax.plot(list(range(len(self.data))),self.data, lw=2, c = 'black')

        if self.save_click : 
            self.list_x = []
            ax.set_label('good')
            # This function allows us to get different information about the last click (on which ax it was, what were the x,y coordinates of click on the ax etc)
            self.fig.canvas.mpl_connect('button_press_event', self.mouse_event)
        
        if self.frames_to_show is not None :
            ylim = self.ylim
            if type(self.frames_to_show) in [np.ndarray, list]: 
                ax.scatter(self.frames_to_show, [ylim[1]-(10*abs(ylim[1]-ylim[0])/100) for _ in range (len(self.frames_to_show))], c = 'red', s = 100)
            elif type(self.frames_to_show) == dict : 
                for color in self.frames_to_show.keys() :
                    ax.scatter(self.frames_to_show[color], [ylim[1]-(10*abs(ylim[1]-ylim[0])/100) for _ in range (len(self.frames_to_show[color]))], c = color, s = 100)
        
        ax.set_ylim(self.ylim)

    def previous(self, event) : 
        if self.t - self.window > 0 : 
            self.t -= self.window 
            self.plot()

    def next(self, event) : 
        if self.t + self.window < len(self.data) : 
            self.t += self.window 
            self.plot()

    
    def mouse_event(self, event):
        # If the user clicked on the plot
        if event.inaxes is not None :
            if event.inaxes.get_label() == 'good' : 
                # We get the x coordinate of the mouse when clicking 
                self.list_x.append(int(event.xdata))


    def on_close(self, event):
            self.pause = False 

    def plot(self) : 
        for ax in self.axes : 
            ax.set_xlim([self.t,self.t+self.window-1]) # update the xlim
            ax.set_xticks = ax.get_xticks()
            self.fig.canvas.draw_idle() # draws the updated plot 


## PICKLE functions


def channel_from_pkl(date_time, file, ext = '.pkl') : 
    '''
    For a file named as whatever_date_time_channel.pkl, finds the channel name 
    '''
    index_channel = file.index(date_time)+len(date_time)+1
    index_ext = file.index(ext)
    channel = file[index_channel:index_ext]
    return channel 

def list_date_time_pkl(date_time) :
    '''
    Look in the peaks folder and get all the files, returns only the one with the right date_time in the name 
    ''' 
    list_analyzed = listdir(spe.peaks)
    list_analyzed = [file for file in list_analyzed if date_time in file]
    return list_analyzed



def save_dict(var, pathfile) : 
    '''
    saves variable var in a picke file based on the path chosen 
    NB : do not forget to put .pkl at the end of the path, if just a name like file.pkl will be put in the working directory
    '''
    with open(pathfile, 'wb') as fp:
        pickle.dump(var, fp)

def read_dict(pathfile) :
    '''
    Returns variable from a picke file based on the path of this file 
    '''
    with open(pathfile, 'rb') as fp:
        var = pickle.load(fp)
    return var 




### MINOR USEFUL FUNCTIONS

def expand_ID(start = start_row) : 
    '''
    This function changes the ID in ID_slice
    '''
    
    wb = op.load_workbook(spe.infos) 
    sheet = wb['conditions'] # Go to the right sheet 
    
    # Finds the columns of interest 
    for column in range (1,sheet.max_column) :
        if sheet.cell(row=1, column=column).value == 'ID' :
            column_ID = column 
        if sheet.cell(row=1, column=column).value in ['slice', 'Slice'] :
            column_slice = column 

    for row in range (start,sheet.max_row + 1) : 
        ID = sheet.cell(row=row, column=column_ID).value # Get the exp ID
        slice = sheet.cell(row=row, column=column_slice).value # Get the slice number
        sheet.cell(row=row, column=column_ID).value = f'{ID}_{slice}' # change the value
    
    # save the file
    wb.save(spe.infos)

def expand_matrix(index_i, lim_i, index_j, lim_j, values, filler = 0) : 
    '''
    Put 0 (default) or other filler value in a matrice to expand it to lim_i and lim_j (used in colormap so that the 0 values are everywhere on the picture for an homogeneous background)
    '''
    if lim_i[0] < index_i[0] :
        index_i = [lim_i[0]] + index_i 
        values = [[filler]*len(index_j)] + values

    if lim_i[1] > index_i[-1] :
        index_i = index_i + [lim_i[1]]
        values = values + [[filler]*len(index_j)]

    if lim_j[0] < index_j[0] :
        index_j = [lim_j[0]] + index_j
        for v in range(len(values)) :
            values[v] = values[v] + [filler]

    if lim_j[1] > index_j[-1] :
        index_j = index_j + [lim_j[1]]
        for v in range(len(values)) :
            values[v] = [filler] + values[v]

    return index_i, index_j, values 

def rounding(list) : 
    '''
    I created this function for making prettier list of values 
    todo This is not optimal. 
    '''

    def power_of_x(x) : 
        power = 1
        while x >= 10:
            x /= 10
            power += 1
        return power


    def find_factor(power) :
        if power < 2 :
            factor = 1 
        elif power == 2 :
            factor = 5
        elif power == 3 : 
            factor = 25
        else : 
            factor = int((10 ** (power-2))/4)

        return factor 
    
    rangelist = abs(max(list) - min(list))
    
    if rangelist < 2 : 
        return [round(x,2) for x in list]
    else :
        new_list = []
        for x in list : 
            if abs(x) < 1 :
                new_list.append(round(x,2))
            elif abs(x) < 2 :
                new_list.append(round(x,1))
            else : 
                power = power_of_x(x)
                f = find_factor(power)
                possibilities = [f*(x//f), f*(x//f + 1)]
                new_list.append(int(possibilities[np.argmin([abs(x-p) for p in possibilities])]))

        return new_list

def div (x, d = 10) : 
    m = x //d
    k = x - d*(m)
    return (m,k)


def check_for_problems() :
    '''
    todo : change my code so I take this into account at the beginning, I discovered that I get my data in the unit defined in the InfoChannel, with the exponent defined there
    I have to multiply my data by a conversion factor.. Weird
    I did this to check if they all have the same units, exponent etc but this is not ideal, change this!!!
    '''

    problems = {}
    infos_interest = {'unit' : {'index' : 6, 'val' : b'V'}, 'exponent' : {'index' : 7, 'val' : -12}, 'step' : {'index' : 8, 'val' : 0}, 
                    'freq' : {'index' : 9, 'val' : 100}, 'factor' : {'index' : 10, 'val' : 59605}}

    paths = choose_files_from_info()
    for path in paths :
        problems[path] = []

        date_time = date_time_from_path(path)
        
        if date_time in spe.specials.keys() :
            path = spe.specials[date_time]['path']

        data = h5py.File(path)
        k = 0 
        while True :
            try :
                data = data['Data']['Recording_0']['AnalogStream'][f'Stream_{k}']
                break
            except :
                k += 1
        # Get the channels names
        data = data['InfoChannel'] 
        problems[path] = [[] for _ in range (len(data))]
        for c, channel in enumerate(data) : 
            for key in infos_interest.keys() : 
                if channel[infos_interest[key]['index']] != infos_interest[key]['val'] :
                    text = f"{key} : {channel[infos_interest[key]['index']]}"
                    print(text)
                    problems[path][c].append(text)

